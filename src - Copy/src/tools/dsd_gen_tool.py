from win32com.client import Dispatch
import pandas as pd
from datetime import datetime
import os
import traceback
import numpy as np
from typing import Dict, List, Optional, Union, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import json

class AscetDatabaseScanner:
    """Original class to scan ASCET database and retrieve implementation information"""
    def __init__(self, version="6.1.4"):
        """Initialize ASCET database scanner"""
        self.ascet = None
        self.db = None
        self.version = version
        self.all_implementations = {}
        self.processed_paths = set()
        self.available_classes = {}  # Store available classes for selective processing
        self.enhanced_import_params = {}  # Store enhanced import parameter data
        
    def connect(self):
        """Connect to ASCET database"""
        try:
            self.ascet = Dispatch(f"Ascet.Ascet.{self.version}")
            self.db = self.ascet.GetCurrentDataBase()
            print(f"Successfully connected to ASCET {self.version}")
            
            if self.db:
                print(f"Database name: {self.db.GetName()}")
                
                # Diagnostic information about GetAllAscetFolders
                print("\nDiagnostic information:")
                get_folders_attr = getattr(self.db, "GetAllAscetFolders", None)
                print(f"GetAllAscetFolders type: {type(get_folders_attr)}")
                
                # Try to inspect what methods are available
                print("\nAvailable methods:")
                for method_name in dir(self.db):
                    if method_name.startswith("Get"):
                        print(f"- {method_name}")
                
            return True
        except Exception as e:
            print(f"Connection failed: {str(e)}")
            return False
            
    def scan_database_structure_internal(self):
        """Quick scan of database structure to find all classes without processing details"""
        try:
            print("\nStarting database structure scan...")
            
            # Reset available classes
            self.available_classes = {}
            
            # Get all ASCET top folders
            try:
                print("\nGetting top-level folders...")
                top_folders = []
                
                # Method 1: Try accessing GetAllAscetFolders properly
                try:
                    # Get the attribute without calling it
                    get_folders_attr = getattr(self.db, "GetAllAscetFolders", None)
                    print(f"GetAllAscetFolders attribute type: {type(get_folders_attr)}")
                    
                    if get_folders_attr is not None:
                        if callable(get_folders_attr):
                            folders = get_folders_attr()  # Call it if it's callable
                        else:
                            folders = get_folders_attr  # Use it directly if it's a property
                        
                        # Process the result appropriately based on its type
                        if isinstance(folders, tuple) or isinstance(folders, list):
                            for folder in folders:
                                if folder:
                                    top_folders.append(folder)
                        elif folders:  # It might be a single object
                            top_folders.append(folders)
                except Exception as e1:
                    print(f"GetAllAscetFolders method failed: {str(e1)}")
                    
                # Method 2: Try alternative methods if the first approach failed
                if not top_folders:
                    try:
                        # Try GetAllFolders instead
                        folders = self.db.GetAllFolders()
                        if folders:
                            for folder in folders:
                                if folder:
                                    top_folders.append(folder)
                    except Exception as e2:
                        print(f"GetAllFolders method failed: {str(e2)}")
                
                if not top_folders:
                    print("No top-level folders found in the database")
                    return False
                    
                # Process the folders we found
                folder_count = 0
                for folder in top_folders:
                    if folder:
                        try:
                            folder_name = folder.GetName()
                            print(f"\nScanning folder structure: {folder_name}")
                            
                            # Get the full path if available
                            folder_path = folder.GetNameWithPath() if hasattr(folder, 'GetNameWithPath') else folder_name
                            self.find_classes(folder, folder_path)
                            folder_count += 1
                        except Exception as folder_error:
                            print(f"Error processing folder: {str(folder_error)}")
                
                print(f"\nCompleted structure scan of {folder_count} top-level folders")
                print(f"Found {len(self.available_classes)} classes in total")
                return True
                    
            except Exception as e:
                print(f"Error in top-level folder scanning: {str(e)}")
                print("Detailed error information:")
                traceback.print_exc()
                return False
                
        except Exception as e:
            print(f"Error scanning database structure: {str(e)}")
            print("Detailed error information:")
            traceback.print_exc()
            return False

    def find_classes(self, folder, folder_path):
        """Recursively find classes in a folder"""
        try:
            if not folder:
                return
                    
            # Check if it's a class component
            if hasattr(folder, 'IsClass') and folder.IsClass():
                print(f"Found Class: {folder_path}")
                self.available_classes[folder_path] = folder
            
            # Process subfolders
            if hasattr(folder, 'GetSubFolders'):
                subfolders = folder.GetSubFolders()
                if subfolders:
                    for subfolder in subfolders:
                        if subfolder:
                            subfolder_name = subfolder.GetName()
                            subfolder_path = f"{folder_path}\\{subfolder_name}"
                            print(f"Entering subfolder: {subfolder_path}")
                            self.find_classes(subfolder, subfolder_path)
            
            # Process Package folder
            if hasattr(folder, 'GetItemInFolder'):
                package_path = f"{folder_path}\\Package"
                package_folder = folder.GetItemInFolder("Package", folder_path)
                if package_folder:
                    print(f"Entering Package folder: {package_path}")
                    self.find_classes(package_folder, package_path)
            
            # Process all database items
            if hasattr(folder, 'GetAllDataBaseItems'):
                # Handle GetAllDataBaseItems as either property or method
                get_items_attr = getattr(folder, 'GetAllDataBaseItems')
                if callable(get_items_attr):
                    items = get_items_attr()
                else:
                    items = get_items_attr
                    
                if items:
                    for item in items:
                        if not item:
                            continue
                            
                        item_name = item.GetName()
                        item_path = f"{folder_path}\\{item_name}"
                        
                        if hasattr(item, 'IsClass') and item.IsClass():
                            print(f"Found Class: {item_path}")
                            self.available_classes[item_path] = item
                        elif hasattr(item, 'IsFolder') and item.IsFolder():
                            print(f"Entering subfolder: {item_path}")
                            self.find_classes(item, item_path)
                                
        except Exception as e:
            print(f"Error finding classes in folder {folder_path}: {str(e)}")

    def process_selected_class(self, class_path):
        """Process only the selected class"""
        try:
            print(f"\nProcessing selected class: {class_path}")
            
            if class_path not in self.available_classes:
                print(f"Error: Class {class_path} not found in available classes")
                return False
                
            class_component = self.available_classes[class_path]
            
            # Process the class implementation
            self.process_class_implementation(class_component, class_path)
            
            print(f"Completed processing class {class_path}")
            return True
            
        except Exception as e:
            print(f"Error processing selected class: {str(e)}")
            traceback.print_exc()
            return False
    def getelement_value(self, element):
        """
        通用的ASCET数据获取函数
        Parameters:
            element: ASCET模型元素
        Returns:
            tuple: (value, error)
        """
        try:
            di = element.GetValue()
            if not di:
                return None, "GetValue() 返回 None"
                
            # 按常见类型优先级尝试
            type_methods = [
                'GetDoubleValue',   # 双精度浮点
                'GetFloatValue',    # 单精度浮点  
                'GetIntegerValue',  # 整数
                'GetBooleanValue',  # 布尔
                'GetStringValue',   # 字符串
                'GetLongValue'      # 长整型
            ]
            
            for method_name in type_methods:
                if hasattr(di, method_name):
                    try:
                        value = getattr(di, method_name)()
                        print(f"成功 使用 {method_name} 获取到: {value}")
                        return value, None
                    except:
                        continue
            return None, "所有数据类型方法都失败"
        except Exception as e:
            return None, f"获取数据失败: {e}"
        
    def scan_database(self):
        """Scan the entire database by processing all folders"""
        try:
            print("\nStarting full database scan...")
            
            # First scan to identify all classes
            if not self.scan_database_structure_internal():
                print("Failed to scan database structure")
                return False
            
            # Now process each class
            print("\nProcessing all identified classes...")
            processed_count = 0
            
            for class_path, class_component in self.available_classes.items():
                try:
                    print(f"\nProcessing class: {class_path}")
                    self.process_class_implementation(class_component, class_path)
                    processed_count += 1
                except Exception as ce:
                    print(f"Error processing class {class_path}: {str(ce)}")
            
            print(f"\nCompleted processing {processed_count} classes")
            print(f"Total classes with implementation data: {len(self.all_implementations)}")
            return True
            
        except Exception as e:
            print(f"Error scanning database: {str(e)}")
            traceback.print_exc()
            return False

    def process_class_implementation(self, class_component, class_path):
        """Process implementation information for a class component with enhanced implementation type handling"""
        try:
            if not class_component:
                print(f"Error: Invalid class component for {class_path}")
                return
                
            # Handle case where GetAllModelElements is a property or tuple, not a method
            if hasattr(class_component, 'GetAllModelElements'):
                get_elements_attr = getattr(class_component, 'GetAllModelElements')
                if callable(get_elements_attr):
                    model_elements = get_elements_attr()
                    print(f"Obtained model elements via method call for class {class_path}")
                else:
                    # If it's not callable, use it directly as a property
                    model_elements = get_elements_attr
                    print(f"Using GetAllModelElements as a property for class {class_path}")
            else:
                print(f"Class {class_path} does not have GetAllModelElements method")
                return
                    
            if not model_elements:
                print(f"No model elements found in class {class_path}")
                return
                
            elements_info = []
            for element in model_elements:
                try:
                    if not element:
                        continue
                        
                    element_name = element.GetName()
                    print(f"Processing element: {element_name}")
                    
                    # Check if it's an implementation type
                    is_impl_type = False
                    if hasattr(element, 'IsImplInfo'):
                        is_impl_type = element.IsImplInfo()
                    
                    element_type = element.GetModelType() if hasattr(element, 'GetModelType') else ''
                    if is_impl_type:
                        element_type = 'Impl'
                    
                    element_kind = self.get_kind(element)
                    
                    # Clean up method arguments and return values names
                    if element_kind == 'Method Argument' and element_name.startswith('calc/'):
                        element_name = element_name[5:]  # Remove 'calc/' prefix
                    elif element_kind == 'Return Value' and element_name.endswith('/return'):
                        element_name = element_name[:-7]  # Remove '/return' suffix
                    
                    # Initialize element data
                    element_data = {
                        'Name': element_name,
                        'Type': element_type,
                        'Scope': self.get_scope(element),
                        'Kind': element_kind,
                        'Min': '---',
                        'Max': '---',
                        'Q': '---',
                        'Formula': '',  # Default empty value for enums
                        'Impl. Type': '---',
                        'Impl. Min': '---',
                        'Impl. Max': '---',
                        'Unit': self.get_unit(element),
                        'Comment': self.get_comment(element),
                        'Default Value': '---',
                        'Calibration': self.get_calibration_status(element),
                    }
                    
                    # Handle special types: log and enum
                    element_type_lower = element_type.lower()
                    if element_type_lower == 'log':
                        # Initialize with common log values
                        element_data.update({
                            'Min': '0',
                            'Max': '1',
                            'Impl. Min': '0',
                            'Impl. Max': '1',
                            'Q': '1',
                            'Formula': 'ident'
                        })
                        
                        # Properly retrieve implementation type from the API
                        # Try to get the actual implementation type from the BoolImpl object
                        if hasattr(element, 'GetImplementation'):
                            try:
                                impl = element.GetImplementation()
                                if impl:
                                    # Check if it's a BoolImpl object
                                    is_bool_impl = False
                                    if hasattr(impl, 'IsBoolImpl'):
                                        is_bool_impl = impl.IsBoolImpl()
                                    
                                    print(f"Implementation for log element {element_name} is BoolImpl: {is_bool_impl}")
                                    
                                    # Try to get implementation type directly from the implementation object
                                    if hasattr(impl, 'GetImplType'):
                                        impl_type = impl.GetImplType()
                                        if impl_type:
                                            print(f"Implementation type from BoolImpl: {impl_type}")
                                            element_data['Impl. Type'] = impl_type
                                    # Fallback to the standard approach with GetImplInfoForValue
                                    elif hasattr(impl, 'GetImplInfoForValue'):
                                        impl_info = impl.GetImplInfoForValue()
                                        if impl_info and hasattr(impl_info, 'GetImplType'):
                                            impl_type = impl_info.GetImplType()
                                            if impl_type:
                                                print(f"Implementation type from ImplInfo: {impl_type}")
                                                element_data['Impl. Type'] = impl_type
                            except Exception as e:
                                print(f"Error retrieving implementation type for log element: {str(e)}")
                        
                        scope = element_data['Scope']
                        kind = element_data['Kind']
                        
                        # 检查是否应该获取Default Value
                        should_get_log_default = (
                            scope in ['Local', 'Imported'] and 
                            kind in ['Parameter', 'Variable', 'Constant']
                        )
                        
                        print(f"  log元素Default Value检查: scope='{scope}', kind='{kind}', should_get={should_get_log_default}")
                        
                        if should_get_log_default:
                            print(f"  开始为 {scope} {kind} (log类型) 获取Default Value: {element_name}")
                            
                           
                            try:
                                di = element.GetValue()
                                if di:
                                    print(f"  成功获取Value对象")
                                    
                                    #
                                    type_methods = [
                                        'GetDoubleValue',   
                                        'GetFloatValue',    
                                        'GetIntegerValue',
                                        'GetBooleanValue',
                                        'GetStringValue',
                                        'GetLongValue'
                                    ]
                                    
                                    default_found = False
                                    for method_name in type_methods:
                                        if hasattr(di, method_name):
                                            try:
                                                value = getattr(di, method_name)()
                                                print(f"     尝试 {method_name}: {value}")
                                                if value is not None:
                                                    element_data['Default Value'] = str(value)
                                                    print(f"    成功！使用 {method_name} 获取到log默认值: {value}")
                                                    default_found = True
                                                    break
                                            except Exception as method_error:
                                                print(f"     {method_name} 失败: {method_error}")
                                                continue
                                        else:
                                            print(f"    没有 {method_name} 方法")
                                    
                                    if not default_found:
                                        # 如果所有方法都失败，使用log类型的合理默认值
                                   
                                        print(f"    所有获取default Value方法失败")
                                else:
                                    print(f"    element.GetValue() 返回 None")
                                    
                                    
                                    
                            except Exception as e:
                                print(f"    获取Default Value异常: {e}")
                                
                        else:
                            print(f"  跳过log元素Default Value获取 (scope={scope}, kind={kind})")
                    elif element_type_lower == 'enum':
                        # Set initial values for enum type
                        element_data.update({
                            'Min': '0',
                            'Q': '1',
                            'Formula': '',  # Empty for enums
                            'Impl. Type': 'uint8'
                        })
                        
                        # Get enum implementation info
                        impl = None
                        impl_info = None
                        if hasattr(element, 'GetImplementation'):
                            impl = element.GetImplementation()
                            if impl and hasattr(impl, 'GetImplInfoForValue'):
                                impl_info = impl.GetImplInfoForValue()
                        
                        # Get enum max value
                        if impl_info:
                            max_value = self.get_enum_max_value(element, impl_info)
                            if max_value != '---':
                                element_data['Max'] = max_value
                                # Fix: Also set Impl. Max and Impl. Min for enums
                                element_data['Impl. Max'] = max_value
                                element_data['Impl. Min'] = element_data['Min']  # Copy Min to Impl. Min
                    
                    # Get implementation info for other types
                    if hasattr(element, 'GetImplementation') and element_type_lower not in ['log', 'enum']:
                        impl = element.GetImplementation()
                        if impl and hasattr(impl, 'GetImplInfoForValue'):
                            impl_info = impl.GetImplInfoForValue()
                            if impl_info:
                                self.update_implementation_info(element_data, element, impl_info)
                                if hasattr(impl_info, 'GetFormulaName'):
                                    formula_name = impl_info.GetFormulaName()
                                    if formula_name:
                                        element_data['Formula'] = formula_name
                    
                    elements_info.append(element_data)
                    print(f"Successfully obtained element info: {element_name}")
                    
                except Exception as e:
                    print(f"Error processing element {element_name if 'element_name' in locals() else 'unknown'}: {str(e)}")
                    continue
                    
                try:
                        if hasattr(element, 'GetDefaultData'):
                            default_data = element.GetDefaultData()
                            if default_data is not None:
                                element_data['DefaultData'] = default_data
                except Exception as e:
                        print(f"Error getting DefaultData for {element_name}: {e}")
                    
            if elements_info:
                self.all_implementations[class_path] = elements_info
                print(f"Found {len(elements_info)} elements in class {class_path}")
            else:
                print(f"No elements found in class {class_path}")
                
        except Exception as e:
            print(f"Error processing class implementation: {str(e)}")
            traceback.print_exc()


    def get_enum_max_value(self, element, impl_info):
        """Get maximum value for enum type using multiple methods"""
        try:
            # Method 1: Try using GetValue and GetMaxSize
            if hasattr(element, 'GetValue'):
                try:
                    value_item = element.GetValue()
                    if value_item and hasattr(value_item, 'GetMaxSize'):
                        max_size = value_item.GetMaxSize()
                        if max_size is not None:
                            return str(max_size - 1)  # Enum starts from 0, so max is size-1
                except Exception:
                    pass

            # Method 2: Try using various range methods
            range_methods = [
                ('GetDoublePhysicalRange', float),
                ('GetFloatPhysicalRange', float),
                ('GetIntegerPhysicalRange', int),
                ('GetLongPhysicalRange', int),
                ('GetDoubleImplRange', float),
                ('GetFloatImplRange', float),
                ('GetIntegerImplRange', int),
                ('GetLongImplRange', int)
            ]

            for method_name, convert_type in range_methods:
                if hasattr(impl_info, method_name):
                    try:
                        method = getattr(impl_info, method_name)
                        range_value = method()
                        if range_value and len(range_value) == 2:
                            max_val = convert_type(range_value[1])
                            return str(max_val)
                    except Exception:
                        continue

            return '---'

        except Exception as e:
            print(f"Error getting enum max value: {str(e)}")
            return '---'

    def get_scope(self, element):
        """Get the element's scope"""
        try:
            if hasattr(element, 'IsMethodReturn') and element.IsMethodReturn():
                if hasattr(element, 'GetName'):
                    name = element.GetName()
                    if '/return' in name:
                        return name.replace('/return', '')
                    return name
                return 'Unknown'
                    
            if hasattr(element, 'GetScope'):
                scope = element.GetScope()
                if hasattr(element, 'IsMethodArgument') and element.IsMethodArgument():
                    if hasattr(element, 'GetName'):
                        return element.GetName().split('/')[0]
                elif hasattr(element, 'IsMethodLocal') and element.IsMethodLocal():
                    if hasattr(element, 'GetName'):
                        return element.GetName().split('/')[0]
                else:
                    scope_map = {
                        'local': 'Local',
                        'imported': 'Imported',
                        'exported': 'Exported',
                        'calc': 'calc'
                    }
                    return scope_map.get(scope.lower(), scope)
        except Exception:
            pass
        return 'Unknown'

        
    def get_kind(self, element):
        """Get the element's kind"""
        try:
            if hasattr(element, 'IsConstant') and element.IsConstant():
                return 'Constant'
            if hasattr(element, 'IsMethodArgument') and element.IsMethodArgument():
                return 'Method Argument'
            if hasattr(element, 'IsMethodReturn') and element.IsMethodReturn():
                return 'Return Value'
            if hasattr(element, 'IsImplementationCast') and element.IsImplementationCast():
                return 'Impl. Cast'
            if hasattr(element, 'IsParameter') and element.IsParameter():
                return 'Parameter'
            if hasattr(element, 'IsVariable') and element.IsVariable():
                return 'Variable'
            if hasattr(element, 'IsReceiveMessage') and element.IsReceiveMessage():
                return 'Receive Message'
            if hasattr(element, 'IsSendMessage') and element.IsSendMessage():
                return 'Send Message'
            if hasattr(element, 'IsSendReceiveMessage') and element.IsSendReceiveMessage():
                return 'Send Receive Message'
            
            return 'Unknown'
        except Exception:
            return 'Unknown'

    def get_unit(self, element):
        """Get the element's unit"""
        try:
            if hasattr(element, 'GetUnit'):
                unit = element.GetUnit()
                return unit if unit else '---'
        except Exception:
            pass
        return '---'

    def get_comment(self, element):
        """Get the element's comment"""
        try:
            if hasattr(element, 'GetComment'):
                comment = element.GetComment()
                return comment if comment else '---'
        except Exception:
            pass
        return '---'

    def get_calibration_status(self, element):
        """Get the element's calibration status with enhanced logic"""
        try:
            # First check direct calibration status
            if hasattr(element, 'IsCalibration'):
                is_calib = element.IsCalibration()
                if is_calib:
                    return 'calibration'
                else:
                    # If not calibration directly, apply additional checks
                    is_complex = hasattr(element, 'IsComplex') and element.IsComplex()
                    is_method_arg = hasattr(element, 'IsMethodArgument') and element.IsMethodArgument()
                    is_method_return = hasattr(element, 'IsMethodReturn') and element.IsMethodReturn()
                    
                    # Check for imported parameters (special handling)
                    if hasattr(element, 'IsImported') and element.IsImported():
                        # Check if it's explicitly marked as calibratable
                        if hasattr(element, 'IsCalibrationAllowed'):
                            return 'calibration' if element.IsCalibrationAllowed() else 'noncalibration'
                        # Do not assume imported parameters are calibratable by default
                        return 'noncalibration'
                    # Special types that are typically not calibratable
                    elif is_complex or is_method_arg or is_method_return:
                        return '---'
                    else:
                        return 'noncalibration'
            
            # If IsCalibration not available, try alternative methods
            if hasattr(element, 'IsCalibrationAllowed'):
                return 'calibration' if element.IsCalibrationAllowed() else 'noncalibration'
            
            # Check if it's complex or method-related (these are not typically calibratable)
            if (hasattr(element, 'IsComplex') and element.IsComplex()) or \
               (hasattr(element, 'IsMethodArgument') and element.IsMethodArgument()) or \
               (hasattr(element, 'IsMethodReturn') and element.IsMethodReturn()):
                return '---'
                
            # For imported parameters, do not assume they are calibratable
            if hasattr(element, 'IsImported') and element.IsImported():
                # Check for explicit calibration indicator
                if hasattr(element, 'IsCalibrationAllowed'):
                    return 'calibration' if element.IsCalibrationAllowed() else 'noncalibration'
                return 'noncalibration'  # Default for imported parameters
            
            # Default case - if we can't determine, mark as unknown
            return 'Unknown'
        except Exception as e:
            print(f"Error determining calibration status: {str(e)}")
            return 'Unknown'
        
    def update_implementation_info(self, element_data, element, impl_info):
        """
        更新实现信息的改进方法，支持更多类型的Default Value获取
        """
        try:
            element_name = element_data['Name']
            element_type = element_data.get('Type', '')
            scope = element_data['Scope']
            kind = element_data['Kind']
            
            print(f"     update_implementation_info: {element_name} ({element_type}, {scope} {kind})")
            
            # 获取实现类型
            if hasattr(impl_info, 'GetImplType'):
                impl_type = impl_info.GetImplType()
                if impl_type == 'int16':
                    impl_type = 'sint16'
                if impl_type:
                    element_data['Impl. Type'] = impl_type
                    print(f"       ✓ 设置Impl. Type: {impl_type}")
            
            # 获取实现范围
            impl_ranges = self.get_implementation_ranges(impl_info)
            element_data.update(impl_ranges)
            
            # 获取Q值
            q_value = self.get_q_value(element, impl_info)
            element_data['Q'] = q_value
            print(f"       ✓ Q值: {q_value}")
            
            # *** Default Value获取逻辑检查 ***
            should_get_default = (
                (scope == 'Imported' and kind == 'Parameter') or 
                (scope == 'Local' and kind in ['Parameter', 'Variable']) or
                (scope == 'Exported' and kind in ['Parameter', 'Variable']) or
                (scope == 'Local' and kind == 'Constant')
            )
            
            print(f"          检查Default Value条件:")
            print(f"          scope='{scope}', kind='{kind}'")
            print(f"          should_get_default: {should_get_default}")
            
            if should_get_default:
                print(f"       满足条件，开始获取Default Value...")
                
                # 默认值获取方法
                default_value = self.get_default_value(element, impl_info)
                
                print(f"       get_default_value 返回: '{default_value}'")
                
                if default_value != '---':
                    element_data['Default Value'] = default_value
                    print(f"       成功设置Default Value: {default_value}")
                else:
                        print(f"       无法获取Default Value")
            else:
                print(f"       不满足Default Value获取条件，跳过")
            
        except Exception as e:
            print(f"       update_implementation_info异常: {str(e)}")

    def get_implementation_ranges(self, impl_info):
        """Get implementation range information"""
        ranges = {
            'Impl. Min': '---',
            'Impl. Max': '---',
            'Min': '---',
            'Max': '---'
        }
        
        try:
            # Try getting ranges by priority
            range_methods = [
                # First try physical ranges
                ('GetDoublePhysicalRange', float, 'Min', 'Max'),
                ('GetFloatPhysicalRange', float, 'Min', 'Max'),
                ('GetIntegerPhysicalRange', int, 'Min', 'Max'),
                ('GetLongPhysicalRange', int, 'Min', 'Max'),
                # Then try implementation ranges
                ('GetDoubleImplRange', float, 'Impl. Min', 'Impl. Max'),
                ('GetFloatImplRange', float, 'Impl. Min', 'Impl. Max'),
                ('GetIntegerImplRange', int, 'Impl. Min', 'Impl. Max'),
                ('GetLongImplRange', int, 'Impl. Min', 'Impl. Max')
            ]

            for method_name, convert_type, min_key, max_key in range_methods:
                if hasattr(impl_info, method_name):
                    try:
                        method = getattr(impl_info, method_name)
                        range_value = method()
                        if range_value and len(range_value) == 2:
                            ranges[min_key] = str(convert_type(range_value[0]))
                            ranges[max_key] = str(convert_type(range_value[1]))
                    except Exception:
                        continue
                    
        except Exception:
            pass
            
        return ranges

    def get_q_value(self, element, impl_info):
        """Get Q value"""
        try:
            # First check quantization value in implementation info
            if hasattr(impl_info, 'GetQuantization'):
                q_value = impl_info.GetQuantization()
                if q_value is not None:
                    return str(q_value)
            
            # Then check Q value in implementation info
            if hasattr(impl_info, 'GetQ'):
                q_value = impl_info.GetQ()
                if q_value is not None:
                    return str(q_value)
            
            # Finally check Q value in element itself
            if hasattr(element, 'GetQ'):
                q_value = element.GetQ()
                if q_value is not None:
                    return str(q_value)
            
            return '---'
            
        except Exception:
            return '---'
    def get_default_value(self, element, impl_info):
        """
        获取参数默认值的改进方法
        Parameters:
            element: ASCET模型元素
            impl_info: 实现信息对象
        Returns:
            str: 默认值或'---'
        """
        try:
            di = element.GetValue()
            if not di:
                return '---'
                
            # 按常见类型优先级尝试（与第三段代码相同的顺序）
            type_methods = [
                'GetDoubleValue',   # 双精度浮点
                'GetFloatValue',    # 单精度浮点  
                'GetIntegerValue',  # 整数
                'GetBooleanValue',  # 布尔
                'GetStringValue',   # 字符串
                'GetLongValue'      # 长整型
            ]
        
            for method_name in type_methods:
                if hasattr(di, method_name):
                    try:
                        value = getattr(di, method_name)()
                        if value is not None:
                            print(f"    ✓ 使用 {method_name} 获取到默认值: {value}")
                            return str(value)
                    except Exception as e:
                        print(f"    • {method_name} 方法失败: {e}")
                        continue
            
            print("    所有数据类型方法都失败")
            return '---'
        
        except Exception as e:
            print(f"    获取默认值过程中发生异常: {e}")
            return '---'
    def export_to_excel(self, output_dir=None):
        """Export data to Excel file with sorted signals and separate imported parameters"""
        if not self.all_implementations:
            print("No implementation information found to export")
            return False
            
        if not output_dir:
            output_dir = "ASCET_Implementations"
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(output_dir, f"ASCET_Class_Implementations_{timestamp}.xlsx")
        
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for class_path, implementations in self.all_implementations.items():
                    if not implementations:
                        continue
                        
                    df = pd.DataFrame(implementations)
                    
                    # Split data into regular signals and imported parameters
                    imported_params = df[
                        (df['Scope'] == 'Imported') & 
                        (df['Kind'] == 'Parameter')
                    ].copy()
                    
                    regular_signals = df[
                        ~((df['Scope'] == 'Imported') & 
                        (df['Kind'] == 'Parameter'))
                    ].copy()
                    
                    # Define kind order priority
                    kind_order = {
                        'Parameter': 1,
                        'Variable': 2,
                        'Method Argument': 3,
                        'Return Value': 4,
                        'Constant': 5,
                        'Receive Message': 6,
                        'Send Message': 7,
                        'Send Receive Message': 8,
                        'Impl. Cast': 9,
                        'Unknown': 10
                    }
                    
                    # Define scope order priority
                    scope_order = {
                        'Local': 1,
                        'Imported': 2,
                        'Exported': 3,
                        'calc': 4,
                        'Unknown': 5
                    }
                    
                    # Process regular signals
                    if not regular_signals.empty:
                        # Create helper columns for sorting
                        regular_signals['kind_order'] = regular_signals['Kind'].map(
                            lambda x: kind_order.get(x, 999))
                        regular_signals['scope_order'] = regular_signals['Scope'].map(
                            lambda x: scope_order.get(x, 999) if x in scope_order else 3)
                        
                        # Sort by priority
                        regular_signals = regular_signals.sort_values(
                            ['kind_order', 'scope_order', 'Name'])
                        
                        # Remove helper columns
                        regular_signals = regular_signals.drop(['kind_order', 'scope_order'], axis=1)
                        
                        # Define column order for regular signals
                        regular_columns = [
                            'Name', 'Type', 'Min', 'Max', 'Q', 'Formula', 
                            'Impl. Type', 'Impl. Min', 'Impl. Max', 'Unit', 
                            'Comment','Scope', 'Kind'
                        ]
                        
                        # Select and reorder columns
                        regular_signals = regular_signals[regular_columns]
                        
                        # Export regular signals
                        sheet_name = os.path.basename(class_path)[:31]
                        regular_signals.to_excel(writer, sheet_name=sheet_name, index=False)
                        print(f"Exported regular signals for class {class_path}")
                    
                    # Process imported parameters
                    if not imported_params.empty:
                        # Define columns for imported parameters
                        imported_columns = [
                            'Name', 'Type', 'Min', 'Max', 'Default Value', 
                            'Calibrable', 'Unit', 'Comment'
                        ]
                        
                        # Prepare imported parameters data
                        imported_params_processed = pd.DataFrame({
                            'Name': imported_params['Name'],
                            'Type': imported_params['Type'],
                            'Min': imported_params['Min'],
                            'Max': imported_params['Max'],
                            'Default Value': imported_params['Default Value'],
                            'Calibrable': imported_params['Calibration'].map(
                                lambda x: 'Yes' if x == 'calibration' else 'No'),
                            'Unit': imported_params['Unit'],
                            'Comment': imported_params['Comment']
                        })
                        
                        # Sort imported parameters by name
                        imported_params_processed = imported_params_processed.sort_values('Name')
                        
                        # Export imported parameters
                        imported_sheet_name = f"{os.path.basename(class_path)[:27]}_Imp"
                        imported_params_processed.to_excel(
                            writer, 
                            sheet_name=imported_sheet_name, 
                            index=False
                        )
                        print(f"Exported imported parameters for class {class_path}")
                        
            print(f"\nExcel file created: {excel_path}")
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {str(e)}")
            traceback.print_exc()
            return False

    def export_selected_class(self, class_path, output_dir=None):
        """Export only a specific class to Excel with enhanced imported parameter information and additional empty columns"""
        if class_path not in self.all_implementations:
            print(f"No implementation information found for class {class_path}")
            return False
            
        if not output_dir:
            output_dir = "ASCET_Implementations"
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(output_dir, f"{os.path.basename(class_path)}_{timestamp}.xlsx")
        
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                implementations = self.all_implementations[class_path]
                df = pd.DataFrame(implementations)
                
                # Split data into regular signals and imported parameters
                imported_params = df[
                    (df['Scope'] == 'Imported') & 
                    (df['Kind'] == 'Parameter')
                ].copy()
                
                regular_signals = df[
                    ~((df['Scope'] == 'Imported') & 
                    (df['Kind'] == 'Parameter'))
                ].copy()
                
                # Define kind order priority
                kind_order = {
                    'Parameter': 1,
                    'Variable': 2,
                    'Method Argument': 3,
                    'Return Value': 4,
                    'Constant': 5,
                    'Receive Message': 6,
                    'Send Message': 7,
                    'Send Receive Message': 8,
                    'Impl. Cast': 9,
                    'Unknown': 10
                }
                
                # Define scope order priority
                scope_order = {
                    'Local': 1,
                    'Imported': 2,
                    'Exported': 3,
                    'calc': 4,
                    'Unknown': 5
                }
                
                # Process regular signals
                if not regular_signals.empty:
                    # Create helper columns for sorting
                    regular_signals['kind_order'] = regular_signals['Kind'].map(
                        lambda x: kind_order.get(x, 999))
                    regular_signals['scope_order'] = regular_signals['Scope'].map(
                        lambda x: scope_order.get(x, 999) if x in scope_order else 3)
                    
                    # Sort by priority
                    regular_signals = regular_signals.sort_values(
                        ['kind_order', 'scope_order', 'Name'])
                    
                    # Remove helper columns
                    regular_signals = regular_signals.drop(['kind_order', 'scope_order'], axis=1)
                    
                    # Define column order for regular signals
                    regular_columns = [
                        'Name', 'Type', 'Min', 'Max', 'Q', 'Formula', 
                        'Impl. Type', 'Impl. Min', 'Impl. Max', 'Unit', 
                        'Comment','Default Value','Scope', 'Kind'
                    ]
                    
                    # Add empty columns
                    regular_signals['Empty1'] = ''
                    regular_signals['Empty2'] = ''
                    regular_signals['Empty3'] = ''
                    
                    # Select and reorder columns
                    regular_signals = regular_signals[regular_columns]
                    
                    # Export regular signals
                    sheet_name = os.path.basename(class_path)[:31]
                    regular_signals.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"Exported regular signals for class {class_path}")
                
                # Process imported parameters with enhanced information
                if not imported_params.empty:
                    # Prepare enhanced imported parameters data
                    import_data = []
                    
                    # For each imported parameter
                    for _, row in imported_params.iterrows():
                        param_name = row['Name']
                        
                        # Get enhanced data if available
                        if hasattr(self, 'enhanced_import_params') and param_name in self.enhanced_import_params:
                            # Use enhanced data from source class
                            enhanced_data = self.enhanced_import_params[param_name]
                            default_value = enhanced_data.get('Default Value', '---')
                            # Create entry with all enhanced fields and empty columns
                            entry = {
                                'Name': param_name,
                                'Type': enhanced_data.get('Type', row['Type']),
                                'Min': enhanced_data.get('Min', row['Min']),
                                'Empty1': '',  # Empty column before Max
                                'Max': enhanced_data.get('Max', row['Max']),
                                'Default Value': default_value,  # Use Q value as Default Value
                                'Calibrable': 'Yes' if enhanced_data.get('Calibration') == 'calibration' else 'No',
                                'Empty2': '',  # First empty column before Unit
                                'Empty3': '',  # Second empty column before Unit
                                'Unit': enhanced_data.get('Unit', row['Unit']),
                                'Comment': enhanced_data.get('Comment', row['Comment']),
                                'Source': 'Enhanced'
                            }
                        else:
                            # Use original data with empty columns
                            entry = {
                                'Name': param_name,
                                'Type': row['Type'],
                                'Min': row['Min'],
                                'Empty1': '',  # Empty column before Max
                                'Max': row['Max'],
                                'Default Value': row['Q'],  # Use Q as Default Value
                                'Calibrable': 'Yes' if row['Calibration'] == 'calibration' else 'No',
                                'Empty2': '',  # First empty column before Unit
                                'Empty3': '',  # Second empty column before Unit
                                'Unit': row['Unit'],
                                'Comment': row['Comment'],
                                'Source': 'Original'
                            }
                        
                        import_data.append(entry)
                    
                    # Create DataFrame from enhanced data
                    imported_params_processed = pd.DataFrame(import_data)
                    
                    # Sort imported parameters by name
                    imported_params_processed = imported_params_processed.sort_values('Name')
                    
                    # Export imported parameters
                    imported_sheet_name = f"{os.path.basename(class_path)[:27]}_Imp"
                    imported_params_processed.to_excel(
                        writer, 
                        sheet_name=imported_sheet_name, 
                        index=False
                    )
                    print(f"Exported imported parameters for class {class_path}")
                    
            print(f"\nExcel file created: {excel_path}")
            return True
            
        except Exception as e:
            print(f"Error exporting selected class: {str(e)}")
            traceback.print_exc()
            return False
            
    def get_imported_parameters(self, class_path):
        """Get imported parameters for a specific class"""
        if class_path not in self.all_implementations:
            print(f"No implementation information found for class {class_path}")
            return []
            
        implementations = self.all_implementations[class_path]
        imported_params = [item for item in implementations 
                         if item['Scope'] == 'Imported' and item['Kind'] == 'Parameter']
        
        return imported_params

    def disconnect(self):
        """Disconnect from ASCET"""
        if self.ascet:
            try:
                self.ascet.DisconnectFromTool()
                print("Disconnected from ASCET")
            except Exception as e:
                print(f"Error disconnecting: {str(e)}")


class AdvancedAscetLookupTableHandler:
    """Original class to handle ASCET lookup tables"""
    def __init__(self, version: str = "6.1.4"):
        self.ascet = None
        self.db = None
        self.version = version

    def connect(self) -> bool:
        """Connect to ASCET application"""
        try:
            print("Connecting to ASCET...")
            self.ascet = Dispatch(f"Ascet.Ascet.{self.version}")
            self.db = self.ascet.GetCurrentDataBase()
            print("Successfully connected to ASCET database")
            return True
        except Exception as e:
            print(f"Connection error: {str(e)}")
            return False

    def get_array_data(self, array_data_obj) -> List:
        """Extract actual data from ArrayData object trying all available methods"""
        try:
            print(f"Array data object type: {type(array_data_obj).__name__}")
            
            # Try all standard collection methods first
            collection_methods = [
                'getFloatFromCollection',
                'getDoubleFromCollection',
                'getIntegerFromCollection',
                'getLongFromCollection'
            ]
            
            for method in collection_methods:
                if hasattr(array_data_obj, method):
                    try:
                        print(f"Trying {method}...")
                        data = getattr(array_data_obj, method)()
                        if data is not None:
                            result = list(data)
                            print(f"SUCCESS with {method}: Got {len(result)} values")
                            print(f"Sample data: {result[:min(5, len(result))]}")
                            return result
                    except Exception as e:
                        print(f"Error with {method}: {str(e)}")
            
            # Try direct value methods
            value_methods = [
                'GetDoubleValue',
                'GetFloatValue',
                'GetIntegerValue',
                'GetLongValue'
            ]
            
            for method in value_methods:
                if hasattr(array_data_obj, method):
                    try:
                        print(f"Trying {method}...")
                        data = getattr(array_data_obj, method)()
                        if data is not None:
                            result = list(data) if hasattr(data, '__iter__') else [data]
                            print(f"SUCCESS with {method}: Got {len(result)} values")
                            print(f"Sample data: {result[:min(5, len(result))]}")
                            return result
                    except Exception as e:
                        print(f"Error with {method}: {str(e)}")
            
            # Try scalar value methods for single values
            scalar_methods = [
                'GetDoubleScalarValue',
                'GetFloatScalarValue',
                'GetIntegerScalarValue',
                'GetLongScalarValue',
                'GetBooleanValue'
            ]
            
            for method in scalar_methods:
                if hasattr(array_data_obj, method):
                    try:
                        print(f"Trying scalar method {method}...")
                        value = getattr(array_data_obj, method)()
                        print(f"SUCCESS with {method}: Got scalar value {value}")
                        return [value]
                    except Exception as e:
                        print(f"Error with scalar method {method}: {str(e)}")
            
            # If all methods fail, try to get a string representation
            if hasattr(array_data_obj, 'GetStringValue'):
                try:
                    print("Trying GetStringValue...")
                    string_value = array_data_obj.GetStringValue()
                    if string_value:
                        print(f"SUCCESS with GetStringValue: {string_value}")
                        # Try to convert string value to numbers if possible
                        try:
                            import re
                            numbers = re.findall(r"[-+]?\d*\.\d+|\d+", string_value)
                            if numbers:
                                float_values = [float(n) for n in numbers]
                                print(f"Extracted {len(float_values)} values from string")
                                return float_values
                        except:
                            pass
                        return [string_value]
                except Exception as e:
                    print(f"Error with GetStringValue: {str(e)}")
            
            # If we reach here, none of the methods worked
            print("Failed to extract data with any method")
            return []
        except Exception as e:
            print(f"Error in get_array_data: {str(e)}")
            return []

    def is_two_d_table(self, element) -> bool:
        """Determine if element is a 2D table"""
        try:
            print(f"Checking if {element.GetName()} is a 2D table...")
            
            # Check if element is a TwoDTable
            if hasattr(element, 'IsTwoDTable'):
                try:
                    is_2d = element.IsTwoDTable()
                    if is_2d:
                        print(f"Element is confirmed as 2D table via IsTwoDTable()")
                        return True
                except Exception as e:
                    print(f"Error checking IsTwoDTable: {str(e)}")
            
            # Try to use GetYDistribution as indicator of 2D table
            if hasattr(element, 'GetYDistribution'):
                try:
                    y_dist = element.GetYDistribution()
                    if y_dist is not None:
                        print(f"Element has Y distribution -> 2D table")
                        return True
                except Exception as e:
                    print(f"Error checking Y distribution: {str(e)}")
            
            # Alternative check: element has methods specific to 2D tables
            if hasattr(element, 'GetMaxYSize'):
                try:
                    y_size = element.GetMaxYSize()
                    if y_size > 0:
                        print(f"Element has Y size {y_size} -> 2D table")
                        return True
                except Exception as e:
                    print(f"Error checking Y size: {str(e)}")
                    
            # Check if element is explicitly a OneDTable
            if hasattr(element, 'IsOneDTable'):
                try:
                    is_1d = element.IsOneDTable()
                    if is_1d:
                        print(f"Element is confirmed as 1D table via IsOneDTable()")
                        return False
                except Exception as e:
                    print(f"Error checking IsOneDTable: {str(e)}")
            
            print("Element appears to be a 1D table")
            return False
        except Exception as e:
            print(f"Error determining table type: {str(e)}")
            return False

    def try_all_distribution_methods(self, obj, prefix=""):
        """Try all possible distribution methods on an object"""
        x_values = []
        methods_tried = []
        
        # List of distribution methods to try
        distribution_methods = [
            'GetDistribution',
            'GetXDistribution',
            'GetYDistribution',
            'GetAxisDistribution',
            'GetRowDistribution',
            'GetColumnDistribution'
        ]
        
        for method_name in distribution_methods:
            if hasattr(obj, method_name):
                methods_tried.append(method_name)
                try:
                    print(f"{prefix}Trying {method_name}...")
                    dist = getattr(obj, method_name)()
                    if dist:
                        print(f"{prefix}{method_name} returned a valid object")
                        # First try to use GetValue method
                        if hasattr(dist, 'GetValue'):
                            dist_data = dist.GetValue()
                            if dist_data:
                                values = self.get_array_data(dist_data)
                                if values:
                                    print(f"{prefix}Successfully got values from {method_name}")
                                    return values
                            else:
                                print(f"{prefix}Distribution value object is null")
                        else:
                            print(f"{prefix}Distribution object has no GetValue method")
                            
                        # If GetValue is not available or failed, try using the distribution object directly
                        print(f"{prefix}Trying to extract values directly from distribution object")
                        values = self.get_array_data(dist)
                        if values:
                            print(f"{prefix}Successfully extracted values directly from distribution object")
                            return values
                        else:
                            print(f"{prefix}Could not extract values directly from distribution object")
                    else:
                        print(f"{prefix}{method_name} returned None")
                except Exception as e:
                    print(f"{prefix}Error with {method_name}: {str(e)}")
        
        if methods_tried:
            print(f"{prefix}Tried distribution methods: {', '.join(methods_tried)}")
        else:
            print(f"{prefix}No distribution methods found")
            
        return x_values

    def try_get_implementation_data(self, element):
        """Try to access implementation object to get distribution data"""
        try:
            if hasattr(element, 'GetImplementation'):
                print("Trying to get implementation object...")
                impl = element.GetImplementation()
                if impl:
                    print("Implementation object found")
                    # Try all distribution methods on the implementation object
                    return self.try_all_distribution_methods(impl, prefix="Implementation: ")
        except Exception as e:
            print(f"Error getting implementation: {str(e)}")
        return []

    def get_1d_table_data(self, element) -> Optional[Dict]:
        """Get data from a 1D lookup table using GetDistribution as per documentation"""
        try:
            element_name = element.GetName()
            print(f"\nProcessing 1D table: {element_name}")
            
            # Get table value object
            max_size = None
            if hasattr(element,'GetMaxSize'):
                try:
                    max_size = element.GetMaxSize()
                    print("Element MaxSize from GetMaxSize:{max_size}")
                except Exception:
                    pass
            if max_size is None and hasattr(element, 'GetMaxXSize'):
                try:
                    max_size = element.GetMaxXSize()
                    print(f"Element MaxSize from GetMaxXSize: {max_size}")
                except Exception:
                    pass

            value_obj = element.GetValue()
            if not value_obj:
                print(f"No value object found for {element_name}")
                return None
            # Try to get size from value object if not already determined
            if max_size is None and hasattr(value_obj, 'GetSize'):
                try:
                    max_size = value_obj.GetSize()
                    print(f"Value object Size: {max_size}")
                except Exception:
                    pass
                    
            if max_size is None and hasattr(value_obj, 'GetMaxSize'):
                try:
                    max_size = value_obj.GetMaxSize()
                    print(f"Value object MaxSize: {max_size}")
                except Exception:
                    pass
            data = {
                'name': element_name,
                'type': '1D',
                'x_values': [],
                'values': [],
                'max_size': max_size
            }

            # Get table size if available
            if hasattr(value_obj, 'GetSize'):
                try:
                    size = value_obj.GetSize()
                    data['size'] = size
                    print(f"1D Table size: {size}")
                except Exception as e:
                    print(f"Error getting size: {str(e)}")

            # For 1D tables, try multiple approaches to get distribution data
            print("\n---------- Getting X Distribution Data ----------")
            
            # Try all distribution methods on the element
            x_values = self.try_all_distribution_methods(element, prefix="Element: ")
            
            # If that fails, try value object
            if not x_values and value_obj:
                x_values = self.try_all_distribution_methods(value_obj, prefix="ValueObject: ")
                
            # If that fails, try implementation object
            if not x_values:
                x_values = self.try_get_implementation_data(element)
                
            # If x_values found, use them
            if x_values:
                data['x_values'] = x_values
            
            print(f"Final X values: {data.get('x_values', [])}")

            # Get value data
            print("\n---------- Getting Value Data ----------")
            if hasattr(value_obj, 'GetValue'):
                print("Calling value_obj.GetValue()...")
                array_data = value_obj.GetValue()
                if array_data:
                    values = self.get_array_data(array_data)
                    if values:
                        data['values'] = values
                else:
                    print("Value data object is null")
            else:
                print("Value object has no GetValue method")
                
                # Try direct value extraction if GetValue is not available
                print("Trying direct value extraction...")
                values = self.get_array_data(value_obj)
                if values:
                    data['values'] = values
            
            print(f"Final values: {data.get('values', [])}")

            # Get interpolation and extrapolation methods if available
            if hasattr(value_obj, 'IsInterpolationLinear'):
                try:
                    data['interpolation'] = 'linear' if value_obj.IsInterpolationLinear() else 'rounded'
                    print(f"Interpolation: {data['interpolation']}")
                except Exception as e:
                    print(f"Error getting interpolation: {str(e)}")
                
            if hasattr(value_obj, 'IsExtrapolationConstant'):
                try:
                    data['extrapolation'] = 'constant' if value_obj.IsExtrapolationConstant() else 'linear'
                    print(f"Extrapolation: {data['extrapolation']}")
                except Exception as e:
                    print(f"Error getting extrapolation: {str(e)}")

            return data

        except Exception as e:
            print(f"Error getting 1D table data for {element.GetName()}: {str(e)}")
            return None

    

    def get_distribution_data(self, element, axis_type="X"):
        """Get distribution data with fallback mechanisms
        
        Args:
            element: The ASCET table element
            axis_type: "X" or "Y" for 2D tables, or None for 1D tables
            
        Returns:
            Distribution object if found, otherwise None
        """
        method_name = f"Get{axis_type}Distribution" if axis_type else "GetDistribution"
        
        # Try direct access first
        distribution = None
        if hasattr(element, method_name):
            try:
                distribution = getattr(element, method_name)()
                if distribution:
                    return distribution
            except Exception as e:
                print(f"Error with direct {method_name}: {str(e)}")
        
        # Try indirect access through value object
        if hasattr(element, 'GetValue'):
            try:
                value_obj = element.GetValue()
                if value_obj and hasattr(value_obj, method_name):
                    distribution = getattr(value_obj, method_name)()
                    if distribution:
                        print(f"Found {axis_type} distribution through value object")
                        return distribution
            except Exception as e:
                print(f"Error with indirect {method_name}: {str(e)}")
        
        # Try implementation object as last resort
        if hasattr(element, 'GetImplementation'):
            try:
                impl = element.GetImplementation()
                if impl and hasattr(impl, method_name):
                    distribution = getattr(impl, method_name)()
                    if distribution:
                        print(f"Found {axis_type} distribution through implementation")
                        return distribution
            except Exception as e:
                print(f"Error with implementation {method_name}: {str(e)}")
        
        return None

    def get_distribution_data(self, element, axis_type="X"):
        """Get distribution data with fallback mechanisms
        
        Args:
            element: The ASCET table element
            axis_type: "X" or "Y" for 2D tables
            
        Returns:
            Distribution object if found, otherwise None
        """
        method_name = f"Get{axis_type}Distribution"
        
        # Try direct access first
        distribution = None
        if hasattr(element, method_name):
            try:
                distribution = getattr(element, method_name)()
                if distribution:
                    print(f"Found {axis_type} distribution directly from element")
                    return distribution
            except Exception as e:
                print(f"Error with direct {method_name}: {str(e)}")
        
        # Try indirect access through value object
        if hasattr(element, 'GetValue'):
            try:
                value_obj = element.GetValue()
                if value_obj and hasattr(value_obj, method_name):
                    distribution = getattr(value_obj, method_name)()
                    if distribution:
                        print(f"Found {axis_type} distribution through value object")
                        return distribution
            except Exception as e:
                print(f"Error with indirect {method_name}: {str(e)}")
        
        # Try implementation object as last resort
        if hasattr(element, 'GetImplementation'):
            try:
                impl = element.GetImplementation()
                if impl and hasattr(impl, method_name):
                    distribution = getattr(impl, method_name)()
                    if distribution:
                        print(f"Found {axis_type} distribution through implementation")
                        return distribution
            except Exception as e:
                print(f"Error with implementation {method_name}: {str(e)}")
        
        # Try alternative distribution methods
        alt_methods = []
        if axis_type == "X":
            alt_methods = ['GetDistribution', 'GetAxisDistribution', 'GetRowDistribution']
        else:  # Y
            alt_methods = ['GetColumnDistribution']
        
        for alt_method in alt_methods:
            if hasattr(element, alt_method):
                try:
                    distribution = getattr(element, alt_method)()
                    if distribution:
                        print(f"Found {axis_type} distribution via alternative method {alt_method}")
                        return distribution
                except Exception as e:
                    print(f"Error with alternative method {alt_method}: {str(e)}")
        
        print(f"No {axis_type} distribution found with any method")
        return None

    def extract_values_from_distribution(self, distribution):
        """Extract actual values from a distribution object
        
        Args:
            distribution: ASCET distribution object
            
        Returns:
            List of values or empty list if extraction fails
        """
        if not distribution:
            return []
        
        print(f"Extracting values from distribution object type: {type(distribution).__name__}")
        
        # First try GetValue method
        if hasattr(distribution, 'GetValue'):
            try:
                value_obj = distribution.GetValue()
                if value_obj:
                    print("Value object found in distribution")
                    # Try various value extraction methods
                    for method in ['GetDoubleValue', 'GetFloatValue', 'GetIntegerValue', 'GetLongValue']:
                        if hasattr(value_obj, method):
                            try:
                                print(f"Trying method {method}...")
                                values = getattr(value_obj, method)()
                                if values is not None:
                                    # Convert tuple to list if needed
                                    if isinstance(values, tuple):
                                        result = list(values)
                                    else:
                                        result = values
                                    print(f"SUCCESS with {method}: {len(result) if hasattr(result, '__len__') else 'scalar'} value(s)")
                                    return result
                            except Exception as e:
                                print(f"Error with {method}: {str(e)}")
            except Exception as e:
                print(f"Error getting value object: {str(e)}")
        
        # Try direct value methods on distribution
        for method in ['GetDoubleValue', 'GetFloatValue', 'GetIntegerValue', 'GetLongValue']:
            if hasattr(distribution, method):
                try:
                    print(f"Trying direct method {method}...")
                    values = getattr(distribution, method)()
                    if values is not None:
                        # Convert tuple to list if needed
                        if isinstance(values, tuple):
                            result = list(values)
                        else:
                            result = values
                        print(f"SUCCESS with direct {method}: {len(result) if hasattr(result, '__len__') else 'scalar'} value(s)")
                        return result
                except Exception as e:
                    print(f"Error with direct {method}: {str(e)}")
        
        # Try collection methods on distribution
        collection_methods = [
            'getDoubleFromCollection', 
            'getFloatFromCollection', 
            'getIntegerFromCollection',
            'getLongFromCollection'
        ]
        
        for method in collection_methods:
            if hasattr(distribution, method):
                try:
                    print(f"Trying collection method {method}...")
                    values = getattr(distribution, method)()
                    if values is not None:
                        # Convert tuple to list if needed
                        if isinstance(values, tuple):
                            result = list(values)
                        else:
                            result = values
                        print(f"SUCCESS with collection method {method}: {len(result)} value(s)")
                        return result
                except Exception as e:
                    print(f"Error with collection method {method}: {str(e)}")
        
        # Last resort: try to use the array_data methods directly
        print("Trying to use get_array_data as last resort...")
        try:
            values = self.get_array_data(distribution)
            if values:
                print(f"SUCCESS with get_array_data: {len(values)} value(s)")
                return values
        except Exception as e:
            print(f"Error with get_array_data: {str(e)}")
        
        print("Failed to extract values with any method")
        return []

    def get_2d_table_data(self, element) -> Optional[Dict]:
        """Get data from a 2D lookup table using enhanced distribution access"""
        try:
            element_name = element.GetName()
            print(f"\nProcessing 2D table: {element_name}")
            
            data = {
                'name': element_name,
                'type': '2D',
                'x_values': [],
                'y_values': [],
                'values': []  # Will be a 2D array (list of lists)
            }

            # Get table sizes
            if hasattr(element, 'GetMaxXSize'):
                try:
                    x_size = element.GetMaxXSize()
                    data['x_size'] = x_size
                    print(f"X size: {x_size}")
                except Exception as e:
                    print(f"Error getting X size: {str(e)}")
                    
            if hasattr(element, 'GetMaxYSize'):
                try:
                    y_size = element.GetMaxYSize()
                    data['y_size'] = y_size
                    print(f"Y size: {y_size}")
                except Exception as e:
                    print(f"Error getting Y size: {str(e)}")

            # Get X distribution data using enhanced approach
            print("\n---------- Getting X Distribution Data ----------")
            x_distribution = self.get_distribution_data(element, "X")
            if x_distribution:
                x_values = self.extract_values_from_distribution(x_distribution)
                if x_values:
                    data['x_values'] = x_values
                    print(f"X values: {x_values[:5]}{'...' if len(x_values) > 5 else ''}")
            
            # Get Y distribution data using enhanced approach
            print("\n---------- Getting Y Distribution Data ----------")
            y_distribution = self.get_distribution_data(element, "Y")
            if y_distribution:
                y_values = self.extract_values_from_distribution(y_distribution)
                if y_values:
                    data['y_values'] = y_values
                    print(f"Y values: {y_values[:5]}{'...' if len(y_values) > 5 else ''}")
            
            # Get the value matrix data
            print("\n---------- Getting Value Data ----------")
            value_obj = element.GetValue()
            if value_obj:
                print("Value object found")
                
                if hasattr(value_obj, 'GetValue'):
                    print("Calling value_obj.GetValue()...")
                    matrix_data = value_obj.GetValue()
                    if matrix_data:
                        # Try different approaches to extract values
                        flat_values = self.get_array_data(matrix_data)
                        if flat_values:
                            print(f"Got values with length: {len(flat_values)}")
                            print(f"Sample values: {flat_values[:min(5, len(flat_values))]}") # Add this for debugging
                            
                            # Check if values are already in 2D format
                            if (isinstance(flat_values, (list, tuple)) and 
                                len(flat_values) > 0 and 
                                isinstance(flat_values[0], (list, tuple))):
                                
                                print("Values are already in 2D format")
                                data['values'] = flat_values
                                
                            # If not in 2D format but we have dimensions, try to reshape
                            elif data.get('x_values') and data.get('y_values'):
                                x_len = len(data['x_values'])
                                y_len = len(data['y_values'])
                                
                                # If it's a collection of rows (common in ASCET)
                                if len(flat_values) == y_len and all(isinstance(item, (list, tuple)) for item in flat_values):
                                    print(f"Values appear to be a collection of {y_len} rows")
                                    data['values'] = flat_values
                                    
                                # If it's a flat array that needs reshaping
                                elif len(flat_values) == x_len * y_len:
                                    try:
                                        # Reshape into 2D matrix
                                        print(f"Reshaping flat array of length {len(flat_values)} into {y_len}×{x_len} matrix")
                                        matrix = np.array(flat_values).reshape(y_len, x_len)
                                        data['values'] = matrix.tolist()
                                        print(f"Successfully reshaped into {y_len}×{x_len} matrix")
                                        print(f"First row sample: {matrix[0][:min(5, x_len)]}")
                                    except Exception as e:
                                        print(f"Error reshaping matrix: {str(e)}")
                                        # Try alternative approach: manually create 2D array
                                        try:
                                            print("Trying manual 2D array creation...")
                                            manual_matrix = []
                                            for i in range(y_len):
                                                row = []
                                                for j in range(x_len):
                                                    index = i * x_len + j
                                                    if index < len(flat_values):
                                                        row.append(flat_values[index])
                                                if row:
                                                    manual_matrix.append(row)
                                            if manual_matrix:
                                                data['values'] = manual_matrix
                                                print(f"Manually created {len(manual_matrix)}×{len(manual_matrix[0])} matrix")
                                            else:
                                                print("Manual matrix creation failed")
                                                data['values'] = flat_values  # Keep flat values as last resort
                                        except Exception as e2:
                                            print(f"Error in manual matrix creation: {str(e2)}")
                                            data['values'] = flat_values  # Keep flat values as last resort
                                else:
                                    print(f"Cannot reshape: data length {len(flat_values)} != x_len*y_len ({x_len}*{y_len}={x_len*y_len})")
                                    # Try to analyze data structure to determine if it has a pattern
                                    try:
                                        print("Analyzing data structure for patterns...")
                                        # Check if it might be transposed
                                        if len(flat_values) == y_len * x_len:
                                            print("Data might be transposed, attempting alternative reshaping")
                                            matrix = np.array(flat_values).reshape(x_len, y_len).T
                                            data['values'] = matrix.tolist()
                                        else:
                                            data['values'] = flat_values
                                    except Exception as e:
                                        print(f"Error analyzing data: {str(e)}")
                                        data['values'] = flat_values
                            else:
                                print("Cannot reshape: missing X or Y values")
                                data['values'] = flat_values
                        else:
                            print("Failed to extract flat values")
                    else:
                        print("Matrix data object is null")
                else:
                    print("Value object has no GetValue method")
                    
                    # Try direct value extraction if GetValue is not available
                    print("Trying direct value extraction...")
                    values = self.get_array_data(value_obj)
                    if values:
                        print(f"Direct extraction succeeded, got {len(values)} values")
                        data['values'] = values

            return data

        except Exception as e:
            print(f"Error getting 2D table data for {element.GetName()}: {str(e)}")
            traceback.print_exc()
            return None

    def extract_cell_value(self, value):
        """Extract a scalar value from potentially complex data structures"""
        # Handle None values
        if value is None:
            return 0  # Or return "" if you prefer empty cells
            
        # Handle scalar values
        if isinstance(value, (int, float, str, bool)):
            return value
            
        # Handle tuple/list structures
        if isinstance(value, (list, tuple)):
            # Empty container
            if len(value) == 0:
                return 0
                
            # Single-item container
            if len(value) == 1:
                return self.extract_cell_value(value[0])
                
            # Nested container (like in 2D tables)
            if isinstance(value[0], (list, tuple)):
                # Try to find a non-zero value in the nested structure
                for item in value:
                    for subitem in item:
                        if subitem != 0 and subitem is not None:
                            return subitem
                return 0
                
            # Regular container with multiple values
            # Try to find first non-zero value
            for item in value:
                if item != 0 and item is not None:
                    return item
            # If all zeros, return first value
            return value[0]
            
        # For any other type, convert to string
        return str(value)

    def create_safe_sheet_name(self, name, used_names=None):
        """Create a safe, unique Excel sheet name under 31 characters"""
        if used_names is None:
            used_names = set()
            
        # Replace invalid characters
        safe_name = name.replace(':', '_').replace('\\', '_').replace('/', '_')
        
        # Truncate to 28 characters to leave room for uniqueness suffix
        if len(safe_name) > 28:
            safe_name = safe_name[:28]
            
        # Ensure uniqueness
        original_name = safe_name
        suffix = 1
        while safe_name in used_names:
            safe_name = f"{original_name[:28-len(str(suffix))]}_{suffix}"
            suffix += 1
            
        used_names.add(safe_name)
        return safe_name

    def save_tables_to_excel(self, tables: List[Dict], filename: str = "lookup_tables.xlsx") -> bool:
        """Save multiple tables to a single Excel file with proper formatting for lookup tables"""
        try:
            if not tables:
                print("No tables to save")
                return True  # Return True instead of False when no tables exist
                
            # Pre-filter tables to check if we'll have any to save
            tables_to_include = []
            
            for table in tables:
                # Skip tables with no data
                if not table or 'name' not in table:
                    print("Skipping table with no name")
                    continue
                
                # Modified filtering for 1D tables
                if table['type'] == '1D':
                    # First check if we have a defined max_size that's greater than 1
                    max_size = table.get('max_size')
                    if max_size is not None and max_size > 1:
                        print(f"Including 1D table {table['name']} with max_size: {max_size}")
                        tables_to_include.append(table)
                    else:
                        # If no max_size or max_size <= 1, check the actual data
                        x_values_insufficient = not table.get('x_values') or len(table.get('x_values', [])) <= 1
                        values_insufficient = not table.get('values') or len(table.get('values', [])) <= 1
                        
                        if x_values_insufficient and values_insufficient:
                            print(f"Skipping 1D table {table['name']} with insufficient data points")
                        else:
                            tables_to_include.append(table)
                # For 2D tables
                else:
                    # Skip 2D tables with missing X or Y values
                    if not table.get('x_values') or not table.get('y_values'):
                        print(f"Skipping 2D table {table['name']} due to missing X or Y values")
                    else:
                        tables_to_include.append(table)
            
            # If no tables will be included, return success without creating workbook
            if not tables_to_include:
                print("No valid tables to include in Excel export")
                return True  # Return True instead of False
                
            # Create a new workbook
            wb = Workbook()
            
            # Remove the default sheet created by openpyxl
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Define styles
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center')
            
            # Track used sheet names to ensure uniqueness
            used_sheet_names = set()
            
            # Process each table
            for table in tables_to_include:
                # Create a sheet for this table with improved name handling
                safe_name = self.create_safe_sheet_name(table['name'], used_sheet_names)
                ws = wb.create_sheet(title=safe_name)
                
                # Add title and metadata
                ws['A1'] = f"Table: {table['name']}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:D1')
                
                ws['A2'] = f"Type: {table['type']}"
                if table['type'] == '1D':
                    size_value = table.get('max_size', table.get('size', 'N/A'))
                    ws['B2'] = f"Size: {size_value}"
                else:
                    ws['B2'] = f"Size: {table.get('x_size', 'N/A')} x {table.get('y_size', 'N/A')}"
                
                # Add interpolation/extrapolation info if available
                row_idx = 3
                if 'interpolation' in table:
                    ws[f'A{row_idx}'] = f"Interpolation: {table['interpolation']}"
                    row_idx += 1
                if 'extrapolation' in table:
                    ws[f'A{row_idx}'] = f"Extrapolation: {table['extrapolation']}"
                    row_idx += 1
                
                # Add a blank row
                row_idx += 1
                
                # Format 1D table - HORIZONTAL LAYOUT
                if table['type'] == '1D':
                    # [1D table formatting code - unchanged]
                    # Check if we have both x_values and values
                    if table.get('x_values') and table.get('values'):
                        # Add "X values" header
                        ws[f'A{row_idx}'] = "X values"
                        ws[f'A{row_idx}'].font = header_font
                        ws[f'A{row_idx}'].fill = header_fill
                        ws[f'A{row_idx}'].border = thin_border
                        ws[f'A{row_idx}'].alignment = center_alignment
                        
                        # Add X values in row - ensure proper conversion
                        for i, x in enumerate(table['x_values']):
                            col = get_column_letter(i + 2)  # Column B, C, D, etc.
                            
                            # Extract scalar value using the helper function
                            x_value = self.extract_cell_value(x)
                                
                            ws[f'{col}{row_idx}'] = x_value
                            ws[f'{col}{row_idx}'].border = thin_border
                            ws[f'{col}{row_idx}'].alignment = center_alignment
                        
                        # Add "Values" header in next row
                        row_idx += 1
                        ws[f'A{row_idx}'] = "Values"
                        ws[f'A{row_idx}'].font = header_font
                        ws[f'A{row_idx}'].fill = header_fill
                        ws[f'A{row_idx}'].border = thin_border
                        ws[f'A{row_idx}'].alignment = center_alignment
                        
                        # Add values in row - ensure proper conversion
                        for i, v in enumerate(table['values']):
                            col = get_column_letter(i + 2)  # Column B, C, D, etc.
                            
                            # Extract scalar value using the helper function
                            cell_value = self.extract_cell_value(v)
                                
                            ws[f'{col}{row_idx}'] = cell_value
                            ws[f'{col}{row_idx}'].border = thin_border
                            ws[f'{col}{row_idx}'].alignment = center_alignment
                    else:
                        # Show whatever we have
                        if table.get('x_values'):
                            ws[f'A{row_idx}'] = "X Values"
                            ws[f'A{row_idx}'].font = header_font
                            ws[f'A{row_idx}'].fill = header_fill
                            
                            # Add X values in row
                            for i, x in enumerate(table['x_values']):
                                col = get_column_letter(i + 2)
                                x_value = self.extract_cell_value(x)
                                ws[f'{col}{row_idx}'] = x_value
                                ws[f'{col}{row_idx}'].border = thin_border
                                ws[f'{col}{row_idx}'].alignment = center_alignment
                        
                        row_idx += 1
                        
                        if table.get('values'):
                            ws[f'A{row_idx}'] = "Values"
                            ws[f'A{row_idx}'].font = header_font
                            ws[f'A{row_idx}'].fill = header_fill
                            
                            # Add values in row
                            for i, v in enumerate(table['values']):
                                col = get_column_letter(i + 2)
                                cell_value = self.extract_cell_value(v)
                                ws[f'{col}{row_idx}'] = cell_value
                                ws[f'{col}{row_idx}'].border = thin_border
                                ws[f'{col}{row_idx}'].alignment = center_alignment
                                
                # Format 2D table with Y axis in the first column and X axis in the header row
                else:
                    # [2D table formatting code - unchanged]
                    # Print debug info about structure
                    print(f"Processing 2D table: {table['name']}")
                    print(f"X values count: {len(table.get('x_values', []))}")
                    print(f"Y values count: {len(table.get('y_values', []))}")
                    if table.get('values') and len(table.get('values')) > 0:
                        print(f"Values rows: {len(table['values'])}")
                        if isinstance(table['values'][0], (list, tuple)):
                            print(f"First row length: {len(table['values'][0])}")
                    
                    # Process data for the table
                    x_values = [self.extract_cell_value(x) for x in table['x_values']]
                    y_values = [self.extract_cell_value(y) for y in table['y_values']]
                    
                    # Y/X label in top-left corner
                    ws[f'A{row_idx}'] = "Y/X"
                    ws[f'A{row_idx}'].font = header_font
                    ws[f'A{row_idx}'].fill = header_fill
                    ws[f'A{row_idx}'].border = thin_border
                    ws[f'A{row_idx}'].alignment = center_alignment
                    
                    # Header row with X values
                    for i, x in enumerate(x_values):
                        col = get_column_letter(i + 2)  # Column B, C, D, etc.
                        ws[f'{col}{row_idx}'] = x
                        ws[f'{col}{row_idx}'].font = header_font
                        ws[f'{col}{row_idx}'].fill = header_fill
                        ws[f'{col}{row_idx}'].border = thin_border
                        ws[f'{col}{row_idx}'].alignment = center_alignment
                    
                    # Get the matrix values - handle different possible formats
                    matrix = table['values']
                    
                    # Process each row of data
                    for i, y_val in enumerate(y_values):
                        data_row = row_idx + i + 1
                        
                        # Y value in first column
                        ws[f'A{data_row}'] = y_val
                        ws[f'A{data_row}'].font = header_font
                        ws[f'A{data_row}'].fill = header_fill
                        ws[f'A{data_row}'].border = thin_border
                        ws[f'A{data_row}'].alignment = center_alignment
                        
                        # Matrix values
                        # Get current row of values based on the matrix structure
                        row_data = None
                        if i < len(matrix):
                            if isinstance(matrix[i], (list, tuple)):
                                row_data = matrix[i]
                            else:
                                # If we have a flat structure, try to extract the right segment
                                if isinstance(matrix, (list, tuple)) and len(x_values) > 0:
                                    start_idx = i * len(x_values)
                                    end_idx = start_idx + len(x_values)
                                    if start_idx < len(matrix) and end_idx <= len(matrix):
                                        row_data = matrix[start_idx:end_idx]
                        
                        # If we have row data, process each cell
                        if row_data:
                            for j, value in enumerate(row_data):
                                if j < len(x_values):  # Make sure we don't exceed the X dimension
                                    col = get_column_letter(j + 2)  # Column B, C, D, etc.
                                    
                                    # Extract the proper cell value - handle nested structures
                                    cell_value = self.extract_cell_value(value)
                                    
                                    # Write to the cell
                                    ws[f'{col}{data_row}'] = cell_value
                                    ws[f'{col}{data_row}'].border = thin_border
                                    ws[f'{col}{data_row}'].alignment = center_alignment
                        else:
                            # No data for this row, fill with empty cells
                            for j in range(len(x_values)):
                                col = get_column_letter(j + 2)
                                ws[f'{col}{data_row}'] = 0  # Default to zero for empty cells
                                ws[f'{col}{data_row}'].border = thin_border
                                ws[f'{col}{data_row}'].alignment = center_alignment
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    
                    for cell in column:
                        # Skip cells that are part of a merged range
                        is_merged = False
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                break
                        
                        if is_merged:
                            continue
                            
                        if column_letter is None:
                            column_letter = cell.column_letter
                            
                        try:
                            if cell.value is not None:
                                cell_str = str(cell.value)
                                if len(cell_str) > max_length:
                                    max_length = len(cell_str)
                        except:
                            pass
                    
                    if column_letter and max_length > 0:
                        adjusted_width = max(8, max_length + 2)  # Minimum width of 8 for better readability
                        ws.column_dimensions[column_letter].width = adjusted_width
            
            # Only try to save if we have at least one sheet
            if len(wb.worksheets) > 0:
                # Save the workbook
                try:
                    wb.save(filename)
                    print(f"Successfully saved {len(tables_to_include)} tables to {filename}")
                    return True
                except Exception as e:
                    print(f"Error saving Excel file: {str(e)}")
                    return False
            else:
                print("No sheets were created, skipping Excel file creation")
                return True  # Return success when no sheets were created
                
        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
            traceback.print_exc()
            return False
    def process_element(self, element) -> Optional[Dict]:
        """Process a single element (1D or 2D table)"""
        try:
            element_name = element.GetName()
            print(f"\n==== Processing element: {element_name} ====")

            # Check if the element has basic lookup table methods
            if not hasattr(element, 'GetValue'):
                print(f"Element {element_name} is not a lookup table (no GetValue method)")
                return None

            # Determine if it's a 1D or 2D table and process accordingly
            if self.is_two_d_table(element):
                return self.get_2d_table_data(element)
            else:
                return self.get_1d_table_data(element)

        except Exception as e:
            print(f"Error processing element {element.GetName()}: {str(e)}")
            return None

    def format_table_output(self, table_data: Dict) -> str:
        """Format table data for display"""
        if not table_data:
            return "No data available\n"
            
        output = f"Table: {table_data['name']} (Type: {table_data['type']})\n"
        
        if table_data['type'] == '1D':
            # Format 1D table
            output += f"Size: {table_data.get('size', 'unknown')}\n"
            
            if table_data.get('x_values') and table_data.get('values'):
                if len(table_data['x_values']) == len(table_data['values']):
                    # Format as a nice table with aligned columns
                    output += "\nX\tValue\n"
                    output += "-" * 20 + "\n"
                    for x, v in zip(table_data['x_values'], table_data['values']):
                        output += f"{x}\t{v}\n"
                else:
                    # Different lengths, show separately
                    output += f"\nX values ({len(table_data['x_values'])}): {table_data['x_values']}\n"
                    output += f"Values ({len(table_data['values'])}): {table_data['values']}\n"
            else:
                # Show whatever we have
                if table_data.get('x_values'):
                    output += f"\nX values ({len(table_data['x_values'])}): {table_data['x_values']}\n"
                else:
                    output += "\nNo X values available\n"
                    
                if table_data.get('values'):
                    output += f"Values ({len(table_data['values'])}): {table_data['values']}\n"
                else:
                    output += "No values available\n"
                
            if 'interpolation' in table_data:
                output += f"\nInterpolation: {table_data['interpolation']}\n"
            if 'extrapolation' in table_data:
                output += f"Extrapolation: {table_data['extrapolation']}\n"
        else:
            # Format 2D table
            output += f"Size: {table_data.get('x_size', 'unknown')} x {table_data.get('y_size', 'unknown')}\n"
            
            if table_data.get('x_values'):
                output += f"\nX values ({len(table_data['x_values'])}): {table_data['x_values']}\n"
            else:
                output += "\nNo X values available\n"
                
            if table_data.get('y_values'):
                output += f"Y values ({len(table_data['y_values'])}): {table_data['y_values']}\n"
            else:
                output += "No Y values available\n"
            
            if table_data.get('values'):
                if isinstance(table_data['values'][0], list):
                    # It's a proper 2D matrix
                    output += "\nMatrix values:\n"
                    
                    # Format header with X values
                    if table_data.get('x_values'):
                        output += "\t" + "\t".join(map(str, table_data['x_values'])) + "\n"
                        
                        # Format rows with Y values and data
                        for i, row in enumerate(table_data['values']):
                            y_val = table_data['y_values'][i] if i < len(table_data.get('y_values', [])) else f"Row {i}"
                            output += f"{y_val}\t" + "\t".join(map(str, row)) + "\n"
                    else:
                        # No X values, just show matrix
                        for row in table_data['values']:
                            output += "\t".join(map(str, row)) + "\n"
                else:
                    # It's a flat array
                    output += f"\nFlat values ({len(table_data['values'])}): {table_data['values'][:min(20, len(table_data['values']))]}"
                    if len(table_data['values']) > 20:
                        output += "...\n"
                    else:
                        output += "\n"
            else:
                output += "\nNo values available\n"
                
        return output

    def save_table_to_csv(self, table_data: Dict, filename: Optional[str] = None) -> bool:
        """Save table data to CSV file"""
        try:
            if not table_data:
                print("No data to save")
                return False
                
            if not filename:
                filename = f"{table_data['name'].replace('::', '_')}.csv"
                
            if table_data['type'] == '1D':
                # Save 1D table
                if table_data.get('x_values') or table_data.get('values'):
                    # Create DataFrame with whatever data we have
                    df_data = {}
                    
                    if table_data.get('x_values'):
                        df_data['X'] = table_data['x_values']
                        
                    if table_data.get('values'):
                        df_data['Value'] = table_data['values']
                        
                        # If we have values but no X values, create index
                        if not table_data.get('x_values'):
                            df_data['Index'] = list(range(len(table_data['values'])))
                    
                    df = pd.DataFrame(df_data)
                    df.to_csv(filename, index=False)
                    print(f"1D table saved to {filename}")
                    return True
                else:
                    print("No data to save")
                    return False
            else:
                # Save 2D table
                if table_data.get('values'):
                    if isinstance(table_data['values'][0], list):
                        # It's a proper 2D matrix
                        columns = table_data.get('x_values', None)
                        index = table_data.get('y_values', None)
                        
                        df = pd.DataFrame(
                            data=table_data['values'],
                            columns=columns,
                            index=index
                        )
                    else:
                        # It's a flat array
                        df = pd.DataFrame({'Values': table_data['values']})
                        
                    df.to_csv(filename)
                    print(f"2D table saved to {filename}")
                    return True
                else:
                    print("No values to save")
                    return False
                    
        except Exception as e:
            print(f"Error saving table to CSV: {str(e)}")
            return False

    def process_class_tables(self, class_path: str) -> List[Dict]:
        """Process all lookup tables in the specified class"""
        try:
            path_parts = class_path.split('\\')
            class_name = path_parts[-1]
            folder_path = '\\'.join(path_parts[:-1])
            
            print(f"Looking for class: {class_path}")
            class_component = self.db.GetItemInFolder(class_name, folder_path)
            if not class_component:
                print(f"Class component not found: {class_path}")
                return []
            
            model_elements = class_component.GetAllModelElements()
            if not model_elements:
                print(f"No model elements found in {class_path}")
                return []
            
            print(f"Found {len(model_elements)} model elements")
            
            tables = []
            for element in model_elements:
                table_data = self.process_element(element)
                if table_data:
                    tables.append(table_data)
            
            print(f"Successfully processed {len(tables)} lookup tables")
            return tables
            
        except Exception as e:
            print(f"Error processing class tables: {str(e)}")
            return []

    def get_table_by_name(self, class_path: str, table_name: str) -> Optional[Dict]:
        """Get specific table by name from a class"""
        try:
            path_parts = class_path.split('\\')
            class_name = path_parts[-1]
            folder_path = '\\'.join(path_parts[:-1])
            
            print(f"Looking for class: {class_path}")
            class_component = self.db.GetItemInFolder(class_name, folder_path)
            if not class_component:
                print(f"Class component not found: {class_path}")
                return None
            
            # Try to get the specific model element by name
            print(f"Looking for table: {table_name}")
            element = class_component.GetModelElement(table_name)
            if not element:
                print(f"Table '{table_name}' not found in {class_path}")
                return None
                
            return self.process_element(element)
            
        except Exception as e:
            print(f"Error getting table by name: {str(e)}")
            return None
    
    def get_multiple_tables(self, class_path: str, table_names: List[str]) -> List[Dict]:
        """Get multiple tables by name from a class"""
        tables = []
        for table_name in table_names:
            table = self.get_table_by_name(class_path, table_name)
            if table:
                tables.append(table)
            else:
                print(f"Warning: Table '{table_name}' not found or could not be processed")
        
        return tables

    def disconnect(self):
        """Disconnect from ASCET"""
        if self.ascet:
            try:
                self.ascet.DisconnectFromTool()
                print("Disconnected from ASCET")
            except Exception as e:
                print(f"Error disconnecting: {str(e)}")


class IntegratedAscetScanner:
    """A combined class that integrates database scanning with lookup table functionality and enumeration handling"""
    def __init__(self, version="6.1.4"):
        # Initialize both original classes
        self.impl_scanner = AscetDatabaseScanner(version)
        self.table_handler = AdvancedAscetLookupTableHandler(version)
        self.version = version
        
        # Store lookup tables found for each class
        self.all_tables = {}
        
        # New storage for enumerations
        self.enum_elements = {}  # Store enum elements found in classes
        self.all_enumerations = {}  # Store all found enumerations
        self.enumeration_details = {}  # Store detailed information about enumerations
        
        # JSON-related attributes from ASCETDataInterface
        self.json_cache = {}  # Cache for JSON data
        self.cached_classes = set()  # Track which classes have been cached
        self.import_sources_map = {}  # Track import sources for each class
        self.enhanced_params_map = {}  # Track enhanced parameters for each class
        
    def connect(self):
        """Connect to ASCET database and share connection between both handlers"""
        if self.impl_scanner.connect():
            # Since we successfully connected with the implementation scanner,
            # share the connection objects with the table handler
            self.table_handler.ascet = self.impl_scanner.ascet
            self.table_handler.db = self.impl_scanner.db
            return True
        return False
    
    def scan_database_structure(self):
        """Scan database structure to identify all available classes"""
        return self.impl_scanner.scan_database_structure_internal()
    
    def check_class_exists(self, class_path):
        """Check if the specified class exists in the database
        
        Args:
            class_path: Path to the class to check
            
        Returns:
            Boolean indicating whether the class exists
        """
        # Normalize path format
        normalized_path = class_path.replace('/', '\\')
        if normalized_path.startswith('\\'):
            normalized_path = normalized_path[1:]
            
        # Check exact match first
        if normalized_path in self.impl_scanner.available_classes:
            print(f"Class {class_path} found in database.")
            return True
        
        # Try fuzzy matching if exact path not found
        print(f"Exact path for {class_path} not found. Trying fuzzy matching...")
        matched_paths = [path for path in self.impl_scanner.available_classes.keys() 
                    if normalized_path.split('\\')[-1].lower() in path.lower()]
        
        if matched_paths:
            print(f"Found similar classes: {matched_paths[0]}")
            return True
        
        return False

    def full_path(self, comp):
        try:
            if comp is None:
                return "Unknown"
            
            # Check GetNameWithPath
            if hasattr(comp, "GetNameWithPath"):
                get_name_with_path = getattr(comp, "GetNameWithPath")
                if callable(get_name_with_path):
                    # It's a method, call it
                    path = get_name_with_path()
                    return path
                else:
                    # It's a property, use its value
                    path = get_name_with_path
                    return path
            
            # Fallback to GetName
            if hasattr(comp, "GetName"):
                get_name = getattr(comp, "GetName")
                if callable(get_name):
                    # It's a method, call it
                    name = get_name()
                    return name
                else:
                    # It's a property, use its value
                    name = get_name
                    return name
            
            return "Unknown"
        except Exception:
            return "Unknown"
    def discover_import_sources(self, main_class_path):
        """Automatically discover classes that export elements imported by the main class
        
        Args:
            main_class_path: Path to the main class
            
        Returns:
            List of paths to classes that export elements imported by the main class
        """
        try:
            print(f"\n{'='*80}")
            print(f"DISCOVERING IMPORT SOURCES FOR: {main_class_path}")
            print(f"{'='*80}")
            
            # Normalize path format
            normalized_path = main_class_path.replace('/', '\\')
            if normalized_path.startswith('\\'):
                normalized_path = normalized_path[1:]
                
            # Make sure we have the database structure scanned
            if not self.impl_scanner.available_classes:
                print("Database structure not scanned yet. Performing scan...")
                if not self.scan_database_structure():
                    print("Failed to scan database structure")
                    return []
            
            # Get the main class object
            main_cls = None
            if normalized_path in self.impl_scanner.available_classes:
                main_cls = self.impl_scanner.available_classes[normalized_path]
            else:
                # Try fuzzy matching
                matched_paths = [path for path in self.impl_scanner.available_classes.keys() 
                            if normalized_path.split('\\')[-1].lower() in path.lower()]
                
                if matched_paths:
                    print(f"Using best match for main class: {matched_paths[0]}")
                    main_cls = self.impl_scanner.available_classes[matched_paths[0]]
                    normalized_path = matched_paths[0]  # Update to matched path
            
            if not main_cls:
                print(f"Main class not found: {main_class_path}")
                return []
            
            # Get all imported elements from the main class
            imported_elems = []
            try:
                print("Getting model elements...")
                # Handle GetAllModelElements as either method or property
                get_elements_attr = getattr(main_cls, 'GetAllModelElements')
                if callable(get_elements_attr):
                    all_elements = get_elements_attr()  # Call it if it's a method
                else:
                    all_elements = get_elements_attr  # Use it directly if it's a property
                
                if not all_elements:
                    print("No model elements found in class")
                    return []
                    
                for elem in all_elements:
                    try:
                        if elem and hasattr(elem, "GetScope"):
                            scope = elem.GetScope()
                            if scope and scope.lower() == "imported":
                                imported_elems.append(elem)
                    except Exception as e:
                        print(f"Error checking element scope: {str(e)}")
                        
                print(f"Found {len(imported_elems)} imported elements")
            except Exception as e:
                print(f"Error getting model elements: {str(e)}")
                return []
            
            if not imported_elems:
                print("No imported elements found in class")
                return []
            
            # Create a set to store unique source paths
            unique_sources = set()
            
            # Track which elements come from which sources
            source_to_elements = {}
            
            print("\n=== Searching for export sources ===")
            
            # Check for self-reference in main class first
            for elem in imported_elems:
                try:
                    elem_name = elem.GetName()
                    if hasattr(main_cls, "ExistsExportForImport") and main_cls.ExistsExportForImport(elem):
                        source_path = self.full_path(main_cls)
                        print(f"✓ {elem_name} - Source: {source_path} (self-reference)")
                        
                        # Add to unique sources
                        unique_sources.add(source_path)
                        
                        # Track which elements come from this source
                        if source_path not in source_to_elements:
                            source_to_elements[source_path] = []
                        source_to_elements[source_path].append(elem_name)
                except Exception as e:
                    print(f"Error checking self-reference: {str(e)}")
            
            # Search in all other classes
            all_classes = list(self.impl_scanner.available_classes.values())
            print(f"Searching in {len(all_classes)} classes...")
            
            for i, elem in enumerate(imported_elems, 1):
                try:
                    elem_name = elem.GetName()
                    print(f"[{i}/{len(imported_elems)}] Looking for source of: {elem_name}")
                    
                    # Skip if already found in main class
                    if any(elem_name in elements for source, elements in source_to_elements.items() 
                          if source == self.full_path(main_cls)):
                        continue
                        
                    # Search in all other classes
                    found = False
                    for cls in all_classes:
                        if cls == main_cls:
                            continue
                            
                        try:
                            if hasattr(cls, "ExistsExportForImport") and cls.ExistsExportForImport(elem):
                                source_path = self.full_path(cls)
                                print(f"  ✓ {elem_name} - Source: {source_path}")
                                
                                # Add to unique sources
                                unique_sources.add(source_path)
                                
                                # Track which elements come from this source
                                if source_path not in source_to_elements:
                                    source_to_elements[source_path] = []
                                source_to_elements[source_path].append(elem_name)
                                
                                found = True
                                #break
                        except Exception:
                            #  errors when checking individual classes
                            pass
                            
                    if not found:
                        print(f"  ✗ {elem_name} - No source found")
                except Exception as e:
                    print(f"  Error processing element: {str(e)}")
            
            # Convert unique sources to a list for import_class_paths
            import_class_paths = list(unique_sources)
            
            # Print unique sources and their elements
            print(f"\n=== Discovered Source Classes ({len(import_class_paths)}) ===")
            for i, source_path in enumerate(import_class_paths, 1):
                elements = source_to_elements.get(source_path, [])
                print(f"{i}. {source_path} ({len(elements)} elements)")
            
            return import_class_paths
                
        except Exception as e:
            print(f"Error discovering import sources: {str(e)}")
            traceback.print_exc()
            return []

    def process_class_and_check_enums(self, class_path):
        """Process a class and check if it contains enum elements or enum references
        
        Args:
            class_path: Path to the class to process
            
        Returns:
            Boolean indicating whether enums were found
        """
        # Get class object
        class_obj = None
        
        # Try exact match first
        normalized_path = class_path.replace('/', '\\')
        if normalized_path.startswith('\\'):
            normalized_path = normalized_path[1:]
            
        if normalized_path in self.impl_scanner.available_classes:
            class_obj = self.impl_scanner.available_classes[normalized_path]
        else:
            # Try fuzzy matching
            matched_paths = [path for path in self.impl_scanner.available_classes.keys() 
                        if normalized_path.split('\\')[-1].lower() in path.lower()]
            
            if matched_paths:
                print(f"Using best match: {matched_paths[0]}")
                class_obj = self.impl_scanner.available_classes[matched_paths[0]]
                class_path = matched_paths[0]  # Update path to matched path
        
        if not class_obj:
            print(f"Class {class_path} not found in database")
            return False
        
        # Check for enum elements or references directly
        has_enums = False
        
        if hasattr(class_obj, 'GetAllModelElements'):
            # Handle GetAllModelElements as either method or property
            get_elements_attr = getattr(class_obj, 'GetAllModelElements')
            if callable(get_elements_attr):
                model_elements = get_elements_attr()  # Call it if it's a method
            else:
                model_elements = get_elements_attr  # Use it directly if it's a property
            
            if model_elements:
                enum_count = 0
                
                for element in model_elements:
                    try:
                        if not element:
                            continue
                            
                        element_name = element.GetName()
                        
                        # Check if it's a direct enum type
                        is_direct_enum = False
                        if hasattr(element, 'GetModelType'):
                            element_type = element.GetModelType()
                            is_direct_enum = element_type.lower() == 'enum'
                        
                        # Check if it's an enum reference using GetRepresentedClass
                        is_enum_reference = False
                        if hasattr(element, 'GetRepresentedClass'):
                            try:
                                represented_class = element.GetRepresentedClass()
                                if represented_class and hasattr(represented_class, 'IsEnumeration'):
                                    is_enum_reference = represented_class.IsEnumeration()
                            except Exception:
                                pass
                        
                        # Count if either method identified an enumeration
                        if is_direct_enum or is_enum_reference:
                            enum_count += 1
                            print(f"Found enum element/reference: {element_name}")
                            
                    except Exception as element_error:
                        print(f"Error checking element: {str(element_error)}")
                
                has_enums = enum_count > 0
                if has_enums:
                    print(f"Found {enum_count} enum elements/references in class {class_path}")
                else:
                    print(f"No enum elements or references found in class {class_path}")
        
        return has_enums

    def scan_folder_for_classes(self, folder_path):
        """Scan a specific folder in the ASCET database to find all classes
        
        Args:
            folder_path: Path to the folder to scan (e.g., "\PlatformLibrary\Package\AEB_AutomaticEmergencyBrake")
            
        Returns:
            Dictionary of classes found in the specified folder (path -> class_component)
        """
        try:
            print(f"\n{'='*80}")
            print(f"SCANNING FOLDER FOR CLASSES: {folder_path}")
            print(f"{'='*80}")
            
            # Normalize path format
            normalized_path = folder_path.replace('/', '\\')
            if normalized_path.startswith('\\'):
                normalized_path = normalized_path[1:]
            
            # Connect to database if not already connected
            if not self.impl_scanner.db:
                if not self.connect():
                    print("Failed to connect to ASCET database")
                    return {}
            
            # Get the folder from the database
            folder_components = normalized_path.split('\\')
            folder_name = folder_components[-1]
            parent_path = '\\'.join(folder_components[:-1])
            
            # For top-level folders, parent_path will be empty
            if parent_path:
                print(f"Looking for folder '{folder_name}' in parent path '{parent_path}'")
                folder_obj = self.impl_scanner.db.GetItemInFolder(folder_name, parent_path)
            else:
                # Try to get top-level folder
                print(f"Looking for top-level folder '{folder_name}'")
                
                # First try direct method
                try:
                    folder_obj = self.impl_scanner.db.GetFolder(folder_name)
                    if folder_obj:
                        print(f"Found top-level folder '{folder_name}' using GetFolder method")
                except Exception:
                    folder_obj = None
                
                # If that fails, try to get all top folders and find the matching one
                if not folder_obj:
                    try:
                        top_folders = self.impl_scanner.db.GetAllAscetFolders()
                        if top_folders:
                            for folder in top_folders:
                                if folder and hasattr(folder, 'GetName') and folder.GetName() == folder_name:
                                    folder_obj = folder
                                    print(f"Found top-level folder '{folder_name}' in GetAllAscetFolders")
                                    break
                    except Exception as e:
                        print(f"Error getting top folders: {str(e)}")
            
            if not folder_obj:
                print(f"Folder '{folder_path}' not found in database")
                return {}
            
            # Check if the object is actually a folder
            if hasattr(folder_obj, 'IsFolder') and not folder_obj.IsFolder():
                print(f"Object at path '{folder_path}' is not a folder")
                return {}
            
            # Store classes found in this folder
            folder_classes = {}
            
            # Process the folder to find classes
            print(f"Scanning folder '{folder_path}' for classes...")
            self._scan_folder_recursive(folder_obj, normalized_path, folder_classes)
            
            print(f"Found {len(folder_classes)} classes in folder '{folder_path}'")
            # Display found classes
            if folder_classes:
                print("\nClasses found in folder:")
                for i, (class_path, _) in enumerate(folder_classes.items(), 1):
                    print(f"  {i}. {class_path}")
            
            return folder_classes
        
        except Exception as e:
            print(f"Error scanning folder for classes: {str(e)}")
            import traceback
            traceback.print_exc()
            return {}

    def _scan_folder_recursive(self, folder, folder_path, found_classes):
        """Recursively scan folder to find classes
        
        Args:
            folder: ASCET folder object
            folder_path: Path string for the current folder
            found_classes: Dictionary to store found classes (modified in-place)
        """
        try:
            if not folder:
                return
            
            # Check if folder itself is a class
            if hasattr(folder, 'IsClass') and folder.IsClass():
                print(f"Found Class: {folder_path}")
                found_classes[folder_path] = folder
            
            # Process all items in the folder
            if hasattr(folder, 'GetAllDataBaseItems'):
                # Handle GetAllDataBaseItems as either property or method
                get_items_attr = getattr(folder, 'GetAllDataBaseItems')
                if callable(get_items_attr):
                    items = get_items_attr()
                else:
                    items = get_items_attr
                    
                if items:
                    for item in items:
                        if not item:
                            continue
                            
                        item_name = item.GetName()
                        item_path = f"{folder_path}\\{item_name}"
                        
                        if hasattr(item, 'IsClass') and item.IsClass():
                            print(f"Found Class: {item_path}")
                            found_classes[item_path] = item
                        elif hasattr(item, 'IsFolder') and item.IsFolder():
                            print(f"Entering subfolder: {item_path}")
                            self._scan_folder_recursive(item, item_path, found_classes)
            
            # Process subfolders
            if hasattr(folder, 'GetSubFolders'):
                subfolders = folder.GetSubFolders()
                if subfolders:
                    for subfolder in subfolders:
                        if subfolder:
                            subfolder_name = subfolder.GetName()
                            subfolder_path = f"{folder_path}\\{subfolder_name}"
                            print(f"Entering subfolder: {subfolder_path}")
                            self._scan_folder_recursive(subfolder, subfolder_path, found_classes)
            
            # Process Package folder (special folder in ASCET)
            if hasattr(folder, 'GetItemInFolder'):
                package_path = f"{folder_path}\\Package"
                package_folder = folder.GetItemInFolder("Package", folder_path)
                if package_folder:
                    print(f"Entering Package folder: {package_path}")
                    self._scan_folder_recursive(package_folder, package_path, found_classes)
                    
        except Exception as e:
            print(f"Error scanning folder {folder_path}: {str(e)}")

    def process_folder_classes(self, folder_path, process_all=False, output_dir=None):
        """Process all classes found in a specific folder
        
        Args:
            folder_path: Path to the folder to scan
            process_all: If True, process all classes; if False, just scan without processing
            output_dir: Directory to save output files
            
        Returns:
            Boolean indicating success or failure
        """
        try:
            # Find all classes in the folder
            folder_classes = self.scan_folder_for_classes(folder_path)
            
            if not folder_classes:
                print(f"No classes found in folder {folder_path}")
                return False
            
            if not process_all:
                print(f"Found {len(folder_classes)} classes in folder {folder_path}")
                print("To process these classes, call process_folder_classes with process_all=True")
                return True
            
            # Process each class found in the folder
            processed_count = 0
            for class_path, class_obj in folder_classes.items():
                try:
                    print(f"\n{'='*80}")
                    print(f"Processing class {processed_count+1}/{len(folder_classes)}: {class_path}")
                    print(f"{'='*80}")
                    
                    # Use the existing method to process each class
                    if self.process_class(class_path):
                        processed_count += 1
                        
                        # Export data for this class if output_dir is provided
                        if output_dir:
                            self.export_class_data(class_path, output_dir=output_dir)
                    
                except Exception as class_error:
                    print(f"Error processing class {class_path}: {str(class_error)}")
            
            print(f"\nSuccessfully processed {processed_count} out of {len(folder_classes)} classes")
            return processed_count > 0
            
        except Exception as e:
            print(f"Error processing folder classes: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def normalize_class_path(self, class_path):
        """Normalize a class path to ensure consistent format
        
        Args:
            class_path: Original class path
            
        Returns:
            Normalized class path with consistent backslash format
        """
        # Replace forward slashes with backslashes
        normalized = class_path.replace('/', '\\')
        
        # Remove leading backslash if present
        if normalized.startswith('\\'):
            normalized = normalized[1:]
            
        return normalized

    def find_best_matching_path(self, partial_path, available_paths=None):
        """Find the best matching path from available paths
        
        Args:
            partial_path: The partial or inexact path to match
            available_paths: List of available paths to search in (default: use self.impl_scanner.available_classes)
            
        Returns:
            Best matching path or None if no match found
        """
        if not partial_path:
            return None
            
        if available_paths is None:
            if hasattr(self, 'impl_scanner') and hasattr(self.impl_scanner, 'available_classes'):
                available_paths = list(self.impl_scanner.available_classes.keys())
            else:
                return None
                
        if not available_paths:
            return None
            
        # First try exact match
        if partial_path in available_paths:
            return partial_path
            
        # Then try case-insensitive match
        lowercase_path = partial_path.lower()
        for path in available_paths:
            if path.lower() == lowercase_path:
                return path
                
        # Try matching just the class name (last component)
        class_name = partial_path.split('\\')[-1].lower()
        matches = [path for path in available_paths if path.split('\\')[-1].lower() == class_name]
        if matches:
            return matches[0]
            
        # Try partial path matching
        matches = [path for path in available_paths if class_name in path.lower()]
        if matches:
            return matches[0]
            
        return None
    
    def process_class_direct(self, class_path, import_class_paths=None):
        """Process a class without rescanning the database"""
        try:
            print(f"\nProcessing class directly: {class_path}")
            if import_class_paths:
                print(f"With import classes: {import_class_paths}")
            
            # Normalize path format to match how scanning stores paths
            normalized_path = class_path.replace('/', '\\')
            # Remove leading backslash if present to match scan format
            if normalized_path.startswith('\\'):
                normalized_path = normalized_path[1:]
            
            # Debug output to verify path format
            print(f"Normalized path for processing: {normalized_path}")
            print(f"Path exists in available_classes: {normalized_path in self.impl_scanner.available_classes}")
            
            # Check if class exists in available classes
            if normalized_path not in self.impl_scanner.available_classes:
                print(f"Error: Class {normalized_path} not found in available classes")
                
                # Print a few examples of available classes for debugging
                if self.impl_scanner.available_classes:
                    print("Examples of available class paths:")
                    for i, path in enumerate(list(self.impl_scanner.available_classes.keys())[:5]):
                        print(f"  {i+1}. {path}")
                    
                return False
            
            # Process the implementation data directly without rescanning
            if not self.impl_scanner.process_selected_class(normalized_path):
                print(f"Failed to process implementation data for class {normalized_path}")
                return self.process_cross_class_import(normalized_path, import_class_paths)
            
            # If import classes are provided, handle cross-class import
            if import_class_paths:
                # Normalize import paths too
                normalized_import_paths = []
                for import_path in import_class_paths:
                    # Apply the same normalization
                    norm_import = import_path.replace('/', '\\')
                    if norm_import.startswith('\\'):
                        norm_import = norm_import[1:]
                    normalized_import_paths.append(norm_import)
                    
                return self.process_cross_class_import(normalized_path, normalized_import_paths)
            
            # Process lookup tables if available
            all_elements = self.impl_scanner.all_implementations.get(normalized_path, [])
            if all_elements:
                # Split into imported parameters and non-imported elements
                non_imported_elements = [item for item in all_elements if item['Scope'] != 'Imported']
                
                # Scan for lookup tables, but only for non-imported elements
                if non_imported_elements:
                    print(f"\nScanning class {normalized_path} for lookup tables...")
                    non_imported_names = [elem['Name'] for elem in non_imported_elements]
                    class_tables = self.table_handler.get_multiple_tables(normalized_path, non_imported_names)
                    
                    if class_tables:
                        print(f"Found {len(class_tables)} lookup tables in class {normalized_path}")
                        self.all_tables[normalized_path] = class_tables
            
            # Process enumerations
            enum_elements = [item for item in all_elements if item['Type'].lower() == 'enum']
            if enum_elements:
                print(f"Found {len(enum_elements)} enum elements in class {normalized_path}")
                self.enum_elements[normalized_path] = enum_elements
                self.process_class_enumerations(normalized_path)
            
            return True
            
        except Exception as e:
            print(f"Error processing class directly: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
        
    def extract_class_enumerations(self, class_path, output_dir=None):
        """Extract and export only the unique enumerations used in the specified class
        
        Args:
            class_path: Path to the class to analyze
            output_dir: Directory to save output files
            
        Returns:
            Boolean indicating success or failure
        """
        try:
            print(f"\nExtracting enumerations from class: {class_path}")
            
            # Get class object
            class_obj = None
            normalized_path = class_path.replace('/', '\\')
            if normalized_path.startswith('\\'):
                normalized_path = normalized_path[1:]
                
            if normalized_path in self.impl_scanner.available_classes:
                class_obj = self.impl_scanner.available_classes[normalized_path]
            else:
                # Try fuzzy matching
                matched_paths = [path for path in self.impl_scanner.available_classes.keys() 
                            if normalized_path.split('\\')[-1].lower() in path.lower()]
                
                if matched_paths:
                    print(f"Using best match: {matched_paths[0]}")
                    class_obj = self.impl_scanner.available_classes[matched_paths[0]]
                    class_path = matched_paths[0]  # Update path to matched path
            
            if not class_obj:
                print(f"Class {class_path} not found in database")
                return False
            
            # Use a dictionary to track unique enumerations
            unique_enums = {}  # Key: enum_name + source_path, Value: enum_data
            duplicates = 0
            
            if hasattr(class_obj, 'GetAllModelElements'):
                # Handle GetAllModelElements as either method or property
                get_elements_attr = getattr(class_obj, 'GetAllModelElements')
                if callable(get_elements_attr):
                    model_elements = get_elements_attr()
                    print(f"Obtained model elements via method call for class {class_path}")
                else:
                    # If it's not callable, use it directly as a property
                    model_elements = get_elements_attr
                    print(f"Using GetAllModelElements as a property for class {class_path}")
                
                if model_elements:
                    print(f"Scanning {len(model_elements)} elements in class {class_path}")
                    
                    for element in model_elements:
                        try:
                            if not element:
                                continue
                                
                            element_name = element.GetName()
                            
                            # Method 1: Check direct enum type
                            is_direct_enum = False
                            if hasattr(element, 'GetModelType'):
                                element_type = element.GetModelType()
                                is_direct_enum = element_type.lower() == 'enum'
                            
                            # Method 2: Check represented class
                            is_enum_reference = False
                            represented_class = None
                            
                            if hasattr(element, 'GetRepresentedClass'):
                                try:
                                    represented_class = element.GetRepresentedClass()
                                    if represented_class and hasattr(represented_class, 'IsEnumeration'):
                                        is_enum_reference = represented_class.IsEnumeration()
                                except Exception as ref_error:
                                    pass
                            
                            # Process if either method identified an enumeration
                            if is_direct_enum or is_enum_reference:
                                print(f"Found enumeration: {element_name}")
                                
                                # Process the appropriate object
                                if represented_class and is_enum_reference:
                                    # Get the path for the represented class
                                    rep_class_path = represented_class.GetNameWithPath() if hasattr(represented_class, 'GetNameWithPath') else represented_class.GetName()
                                    rep_class_name = represented_class.GetName()
                                    
                                    print(f"Processing enumeration class: {rep_class_name} from {rep_class_path}")
                                    enum_data = self.process_enumeration(represented_class, rep_class_path, rep_class_name)
                                    source_path = rep_class_path  # Use the represented class path
                                else:
                                    # Process the element directly
                                    enum_data = self.process_enumeration(element, class_path, element_name)
                                    source_path = class_path  # Use the current class path
                                
                                if enum_data:
                                    # Create a unique key for this enumeration
                                    unique_key = f"{enum_data['name']}_{source_path}"
                                    
                                    # Check if we've already processed this enumeration
                                    if unique_key in unique_enums:
                                        duplicates += 1
                                        print(f"Skipping duplicate enumeration: {enum_data['name']}")
                                    else:
                                        unique_enums[unique_key] = enum_data
                                        print(f"Added unique enumeration: {enum_data['name']}")
                        
                        except Exception as element_error:
                            print(f"Error processing element {element_name}: {str(element_error)}")
            else:
                print(f"Class {class_path} does not have GetAllModelElements method")
                return False
            
            # Convert dictionary values to list for export
            class_enums = list(unique_enums.values())
            
            # Export the results
            if class_enums:
                print(f"Found {len(class_enums)} unique enumerations in class {class_path}")
                if duplicates > 0:
                    print(f"Skipped {duplicates} duplicate enumerations")
                
                if not output_dir:
                    output_dir = "ASCET_Class_Enumerations"
                os.makedirs(output_dir, exist_ok=True)
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_filename = os.path.basename(class_path)
                
                excel_path = os.path.join(output_dir, f"{base_filename}_Enumerations_{timestamp}.xlsx")
                excel_result = self.export_enumerations_to_excel(class_enums, output_dir)
                
                if excel_result:
                    print(f"Successfully exported {len(class_enums)} unique enumerations to Excel: {excel_path}")
                    return True
                else:
                    print("Failed to export enumerations to Excel")
                    return False
            else:
                print(f"No enumerations found in class {class_path}")
                return False
                
        except Exception as e:
            print(f"Error extracting class enumerations: {str(e)}")
            traceback.print_exc()
            return False
    
    def scan_database_for_all_enumerations(self, output_dir=None):
        """Scan the entire database for all enum elements using GetRepresentedClass method
        
        Args:
            output_dir: Optional directory for saving the Excel output
            
        Returns:
            Boolean indicating success or failure
        """
        try:
            print("\n" + "="*80)
            print("STARTING DATABASE-WIDE ENUMERATION SCAN")
            print("="*80)
            
            # Ensure we have the database structure
            if not self.impl_scanner.available_classes:
                print("Database structure not scanned yet. Performing scan...")
                if not self.scan_database_structure():
                    print("Failed to scan database structure")
                    return False
            
            all_enum_data = []
            total_classes = len(self.impl_scanner.available_classes)
            processed_count = 0
            found_enums_count = 0
            
            print(f"\nScanning {total_classes} classes for enumerations...")
            
            # Process each class in the database
            for class_path, class_obj in self.impl_scanner.available_classes.items():
                processed_count += 1
                
                # Progress indicator every 10 classes
                if processed_count % 10 == 0 or processed_count == total_classes:
                    progress = (processed_count / total_classes) * 100
                    print(f"Progress: {processed_count}/{total_classes} classes ({progress:.1f}%)")
                
                try:
                    # Check if class has model elements
                    if hasattr(class_obj, 'GetAllModelElements'):
                        model_elements = class_obj.GetAllModelElements()
                        
                        if model_elements:
                            # Look for potential enumeration elements
                            for element in model_elements:
                                try:
                                    if not element:
                                        continue
                                        
                                    element_name = element.GetName()
                                    
                                    # Method 1: Check if element has GetRepresentedClass method 
                                    # and its represented class is an enumeration
                                    is_enum_reference = False
                                    represented_class = None
                                    
                                    if hasattr(element, 'GetRepresentedClass'):
                                        try:
                                            represented_class = element.GetRepresentedClass()
                                            if represented_class and hasattr(represented_class, 'IsEnumeration'):
                                                is_enum_reference = represented_class.IsEnumeration()
                                                if is_enum_reference:
                                                    print(f"\nFound enumeration reference: {element_name} in {class_path}")
                                                    print(f"  References enumeration class: {represented_class.GetName()}")
                                        except Exception as ref_error:
                                            # Failed to get represented class, continue with other checks
                                            pass
                                    
                                    # Method 2: Direct type check (original method as fallback)
                                    is_direct_enum = False
                                    if hasattr(element, 'GetModelType'):
                                        element_type = element.GetModelType()
                                        is_direct_enum = element_type.lower() == 'enum'
                                        
                                    # Process if either method identified an enumeration
                                    if is_enum_reference or is_direct_enum:
                                        # If we have a represented class, process that instead
                                        # as it contains the complete enumeration definition
                                        if represented_class:
                                            # Get the path for the represented class
                                            rep_class_path = represented_class.GetNameWithPath() if hasattr(represented_class, 'GetNameWithPath') else represented_class.GetName()
                                            rep_class_name = represented_class.GetName()
                                            
                                            print(f"Processing enumeration class: {rep_class_name}")
                                            enum_data = self.process_enumeration(represented_class, rep_class_path, rep_class_name)
                                        else:
                                            # Process the element directly
                                            enum_data = self.process_enumeration(element, class_path, element_name)
                                        
                                        if enum_data:
                                            all_enum_data.append(enum_data)
                                            # Store for future reference with unique key
                                            key = f"{enum_data['path']}\\{enum_data['name']}"
                                            self.enumeration_details[key] = enum_data
                                            found_enums_count += 1
                                            print(f"  Successfully processed enumeration: {enum_data['name']}")
                                except Exception as element_error:
                                    print(f"Error processing element {element_name}: {str(element_error)}")
                except Exception as class_error:
                    print(f"Error processing class {class_path}: {str(class_error)}")
            
            print(f"\nCompleted database scan: Processed {processed_count} classes, found {found_enums_count} enumerations")
            
            # Export the results if any enumerations were found
            if all_enum_data:
                if not output_dir:
                    output_dir = "ASCET_Enumerations_All"
                os.makedirs(output_dir, exist_ok=True)
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                excel_result = self.export_enumerations_to_excel(all_enum_data, output_dir)
                
                if excel_result:
                    print(f"\nSuccessfully exported {found_enums_count} enumerations to Excel")
                else:
                    print("\nFailed to export enumerations to Excel")
            else:
                print("\nNo enumerations found in the database")
            
            return found_enums_count > 0
                
        except Exception as e:
            print(f"Error scanning database for enumerations: {str(e)}")
            traceback.print_exc()
            return False
    
    def process_class(self, class_path, import_class_path=None):
        """Process a class for implementation data and lookup tables
        
        Args:
            class_path: Path to the main class
            import_class_path: Optional path to import source class. If None, sources are auto-discovered.
            
        Returns:
            Boolean indicating success or failure
        """
        try:
            print(f"\n{'='*80}")
            print(f"Processing main class: {class_path}")
            if import_class_path:
                print(f"Import source class: {import_class_path}")
            print(f"{'='*80}")
            
            # If no import_class_path provided, automatically discover import sources
            if import_class_path is None:
                print("No import source specified. Automatically discovering import sources...")
                import_class_paths = self.discover_import_sources(class_path)
                if import_class_paths:
                    print(f"Discovered {len(import_class_paths)} import sources")
                else:
                    print("No import sources discovered")
                
                # Continue with discovered import sources
                return self.process_cross_class_import(class_path, import_class_paths)
            
            # If import_class_path is provided, use it as a list
            if isinstance(import_class_path, str):
                import_class_paths = [import_class_path]
            else:
                import_class_paths = import_class_path
            
            # Process with the provided import class paths
            return self.process_cross_class_import(class_path, import_class_paths)
            
        except Exception as e:
            print(f"Error processing class: {str(e)}")
            traceback.print_exc()
            return False
    
    def export_class_data(self, class_path, lookup_source_paths=None, output_dir=None):
        """Export both implementation data, lookup tables, and enumerations for a class"""
        try:
            if not output_dir:
                output_dir = "ASCET_Data"
            os.makedirs(output_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = os.path.basename(class_path)
            
            # Export implementation data
            impl_success = self.impl_scanner.export_selected_class(class_path, output_dir)
            
            # Export lookup tables if any
            table_success = True
            if class_path in self.all_tables and self.all_tables[class_path]:
                tables = self.all_tables[class_path]
                
                # Create a suffix for multiple sources if needed
                source_suffix = ""
                if lookup_source_paths:
                    if isinstance(lookup_source_paths, str):
                        lookup_source_paths = [lookup_source_paths]
                    
                    if len(lookup_source_paths) == 1:
                        source_suffix = f"_from_{os.path.basename(lookup_source_paths[0])}"
                    else:
                        source_suffix = f"_from_{len(lookup_source_paths)}_sources"
                
                tables_filename = os.path.join(output_dir, 
                                            f"{base_filename}{source_suffix}_LookupTables_{timestamp}.xlsx")
                
                try:
                    table_success = self.table_handler.save_tables_to_excel(tables, tables_filename)
                    
                    if table_success:
                        print(f"Lookup tables exported to {tables_filename}")
                except Exception as e:
                    error_str = str(e)
                    print(f"Error saving lookup tables to Excel: {error_str}")
                    traceback.print_exc()
                    
                    # Special handling for "At least one sheet must be visible" error
                    # This is not a real failure - it just means there were no tables to export
                    if "At least one sheet must be visible" in error_str:
                        print("This error occurs when there are no valid tables to export - treating as success")
                        table_success = True
                    else:
                        # Try to save to CSV as fallback for other errors
                        print("Trying to save lookup tables to individual CSV files...")
                        csv_dir = os.path.join(output_dir, f"{base_filename}_tables_{timestamp}")
                        os.makedirs(csv_dir, exist_ok=True)
                        
                        csv_save_success = False
                        for i, table in enumerate(tables):
                            if 'name' in table:
                                csv_name = f"{table['name']}.csv"
                            else:
                                csv_name = f"table_{i+1}.csv"
                            
                            csv_path = os.path.join(csv_dir, csv_name)
                            try:
                                self.table_handler.save_table_to_csv(table, csv_path)
                                print(f"Saved table to {csv_path}")
                                csv_save_success = True
                            except Exception as csv_e:
                                print(f"Error saving table to CSV: {str(csv_e)}")
                        
                        # Set table_success based on CSV fallback results
                        table_success = csv_save_success
            else:
                print("No lookup tables to export")
            
            # Export enumerations if any
            enum_success = True
            if class_path in self.enum_elements and self.enum_elements[class_path]:
                print("\nExporting enumerations for class...")
                # Process enumerations if not already done
                if not any(enum_path.startswith(class_path) for enum_path in self.enumeration_details.keys()):
                    self.process_class_enumerations(class_path)
                
                # Get enumeration data for this class
                class_enums = []
                for enum_path, enum_data in self.enumeration_details.items():
                    if enum_data['path'] == class_path:
                        class_enums.append(enum_data)
                
                if class_enums:
                    enums_filename = os.path.join(output_dir, 
                                        f"{base_filename}_Enumerations_{timestamp}.xlsx")
                    enum_success = self.export_enumerations_to_excel(class_enums, output_dir)
                    if enum_success:
                        print(f"Enumerations exported to {enums_filename}")
                else:
                    print("No enumeration details found for export")
            else:
                print("No enumerations to export")
                
            # NEW: Export JSON data - 确保不会导致循环调用
            json_success = True
            print("\nExporting JSON data for class...")
            try:
                # 检查是否已经有JSON缓存数据
                if class_path not in self.json_cache:
                    # 如果没有缓存但有实现数据，直接准备JSON
                    if class_path in self.impl_scanner.all_implementations:
                        print("Preparing JSON from existing implementation data...")
                        # Store enhanced parameters for JSON generation
                        if hasattr(self.impl_scanner, 'enhanced_import_params'):
                            self.enhanced_params_map[class_path] = self.impl_scanner.enhanced_import_params.copy()
                        
                        # Store import sources if provided
                        if lookup_source_paths:
                            self.import_sources_map[class_path] = lookup_source_paths
                        
                        # Mark as cached and prepare JSON
                        self.cached_classes.add(class_path)
                        self._prepare_json_data(class_path)
                    else:
                        print("No implementation data available for JSON export")
                        json_success = False
                
                # Save JSON to file if we have data
                if class_path in self.json_cache:
                    json_filename = os.path.join(output_dir, f"{base_filename}_Data_{timestamp}.json")
                    json_success = self.save_json_to_file(class_path, json_filename)
                    
                    if json_success:
                        print(f"JSON data exported to {json_filename}")
                
            except Exception as json_e:
                print(f"Error exporting JSON data: {str(json_e)}")
                json_success = False
            
            return impl_success and table_success and enum_success and json_success
            
        except Exception as e:
            print(f"Error exporting class data: {str(e)}")
            traceback.print_exc()
            return False
        
    def process_cross_class_import(self, main_class_path, import_class_paths=None):
        """Process a main class and multiple import classes containing its parameters/lookup tables/enumerations
        
        Args:
            main_class_path: Path to the main class
            import_class_paths: Optional list of paths to import source classes. If None, sources are auto-discovered.
            
        Returns:
            Boolean indicating success or failure
        """
        try:
            print(f"\n{'='*100}")
            print(f"Processing cross-class import relationship")
            print(f"Main class: {main_class_path}")
            
            # Auto-discover import sources if none provided
            if import_class_paths is None:
                import_class_paths = self.discover_import_sources(main_class_path)
                if not import_class_paths:
                    print("No import sources found. Processing main class only.")
            
            # Convert to list if a single string is provided
            if isinstance(import_class_paths, str):
                import_class_paths = [import_class_paths]
                
            print(f"Import source classes ({len(import_class_paths)}):")
            for path in import_class_paths:
                print(f"  - {path}")
            print(f"{'='*100}")
            
            # Normalize main path format
            main_path = main_class_path.replace('/', '\\')
            if main_path.startswith('\\'):
                main_path = main_path[1:]  # Remove leading backslash
            
            # Normalize import paths
            normalized_import_paths = []
            for import_path in import_class_paths:
                path = import_path.replace('/', '\\')
                if path.startswith('\\'):
                    path = path[1:]  # Remove leading backslash
                normalized_import_paths.append(path)
            
            # Scan database structure if not done already
            if not self.impl_scanner.available_classes:
                self.scan_database_structure()
            
            # Find matching main class path
            main_exists = main_path in self.impl_scanner.available_classes
            if not main_exists:
                # Try to find similar paths
                matched_paths = [path for path in self.impl_scanner.available_classes.keys() 
                            if main_path.lower() in path.lower()]
                
                if matched_paths:
                    print(f"Exact main class path not found, but found {len(matched_paths)} similar paths:")
                    for i, path in enumerate(matched_paths, 1):
                        print(f"  {i}. {path}")
                    
                    # Use the first match
                    if len(matched_paths) > 0:
                        main_path = matched_paths[0]
                        print(f"Using best match for main class: {main_path}")
                        main_exists = True
                else:
                    print(f"Main class not found: {main_class_path}")
                    return False
            
            # Process the classes with appropriate relationship
            if main_exists:
                # Process the main class first
                if not self.impl_scanner.process_selected_class(main_path):
                    print(f"Failed to process implementation data for class {main_path}")
                    return False
                
                # Get all elements from the implementation data
                all_elements = self.impl_scanner.all_implementations.get(main_path, [])
                
                # Split into imported parameters and non-imported elements
                imported_params = [item for item in all_elements if item['Scope'] == 'Imported' and item['Kind'] == 'Parameter']
                non_imported_elements = [item for item in all_elements if item['Scope'] != 'Imported']
                
                # NEW: Identify enum type elements
                enum_elements = [item for item in all_elements if item['Type'].lower() == 'enum']
                if enum_elements:
                    print(f"\nFound {len(enum_elements)} enum type elements in class {main_path}")
                    for elem in enum_elements:
                        print(f"  - {elem['Name']} (Scope: {elem['Scope']}, Kind: {elem['Kind']})")
                    self.enum_elements[main_path] = enum_elements
                
                all_found_tables = []
                
                # First scan the main class itself for lookup tables, but only for non-imported elements
                if non_imported_elements:
                    print(f"\nScanning main class {main_path} for lookup tables (non-imported elements only)...")
                    non_imported_names = [elem['Name'] for elem in non_imported_elements]
                    main_class_tables = self.table_handler.get_multiple_tables(main_path, non_imported_names)
                    
                    if main_class_tables:
                        print(f"Found {len(main_class_tables)} lookup tables for non-imported elements in main class {main_path}")
                        for table in main_class_tables:
                            print(f"  - {table['name']}")
                        all_found_tables.extend(main_class_tables)
                
                # Handle imported parameters
                if imported_params:
                    print(f"\nFound {len(imported_params)} imported parameters")
                    param_names = [param['Name'] for param in imported_params]
                    
                    # Store parameter sources and enhanced parameter data
                    param_sources = {}  # Track which source provided each parameter
                    enhanced_import_params = {}
                    
                    # Process each import source class
                    for idx, import_path in enumerate(normalized_import_paths):
                        print(f"\n{'='*80}")
                        print(f"Processing import source {idx+1}/{len(normalized_import_paths)}: {import_path}")
                        print(f"{'='*80}")
                        
                        # Find class in database
                        import_exists = import_path in self.impl_scanner.available_classes
                        if not import_exists:
                            # Try to find similar paths
                            matched_paths = [path for path in self.impl_scanner.available_classes.keys() 
                                        if import_path.lower() in path.lower()]
                            
                            if matched_paths:
                                print(f"Exact import path not found, but found {len(matched_paths)} similar paths:")
                                for i, path in enumerate(matched_paths, 1):
                                    print(f"  {i}. {path}")
                                
                                # Use the first match
                                if len(matched_paths) > 0:
                                    import_path = matched_paths[0]
                                    print(f"Using best match for import class: {import_path}")
                                    import_exists = True
                            else:
                                print(f"Import class not found: {import_path}")
                                continue  # Skip this import path and try the next one
                        
                        if import_exists:
                            print(f"Searching for parameters in {import_path}...")
                            
                            # Process the import class to get implementation data
                            if import_path not in self.impl_scanner.all_implementations:
                                self.impl_scanner.process_selected_class(import_path)
                            
                            # Get the implementation data
                            import_impl_data = self.impl_scanner.all_implementations.get(import_path, [])
                            
                            # Try to find lookup tables and implementation information for each parameter
                            for param_name in param_names:
                                # Skip parameters we've already found
                                if param_name in param_sources:
                                    continue
                                    
                                # First, try to get as a lookup table
                                table = self.table_handler.get_table_by_name(import_path, param_name)
                                
                                if table:
                                    # It's a lookup table
                                    print(f"Found lookup table for parameter: {param_name}")
                                    all_found_tables.append(table)
                                    param_sources[param_name] = {
                                        'source': import_path,
                                        'type': 'lookup_table'
                                    }
                                    
                                    # Find implementation data for this parameter
                                    param_impl = next((item for item in import_impl_data if item['Name'] == param_name), None)
                                    if param_impl:
                                        # Make sure we have the Default Value properly set 
                                        if param_impl['Default Value'] == '---' and param_impl['Q'] != '---':
                                            # Use Q value as Default Value if Default Value is not set
                                            param_impl['Default Value'] = param_impl['Q']
                                            print(f"  - Using Q value {param_impl['Q']} as Default Value for {param_name}")
                                        
                                        enhanced_import_params[param_name] = param_impl
                                        print(f"  - Default Value for {param_name}: {param_impl['Default Value']}")
                                else:
                                    # Not a lookup table, try to get implementation info
                                    print(f"Trying to get implementation info for: {param_name}")
                                    
                                    # Find the parameter in the implementations
                                    param_impl = next((item for item in import_impl_data if item['Name'] == param_name), None)
                                    
                                    if param_impl:
                                        print(f"Found implementation info for parameter: {param_name}")
                                        
                                        # Make sure we have the Default Value properly set
                                        if param_impl['Default Value'] == '---' and param_impl['Q'] != '---':
                                            # Use Q value as Default Value if Default Value is not set
                                            param_impl['Default Value'] = param_impl['Q']
                                            print(f"  - Using Q value {param_impl['Q']} as Default Value for {param_name}")
                                        
                                        param_sources[param_name] = {
                                            'source': import_path,
                                            'type': 'implementation',
                                            'data': param_impl
                                        }
                                        enhanced_import_params[param_name] = param_impl
                                        print(f"  - Default Value for {param_name}: {param_impl['Default Value']}")
                                    else:
                                        print(f"Parameter {param_name} not found in {import_path}")
                    
                    # Store found tables
                    if all_found_tables:
                        print(f"\nFound {len(all_found_tables)} lookup tables in total")
                        self.all_tables[main_path] = all_found_tables
                    
                    # Print summary of parameter sources
                    print("\nParameter sources summary:")
                    for param_name in param_names:
                        if param_name in param_sources:
                            source_info = param_sources[param_name]
                            source_type = source_info['type']
                            source_path = source_info['source']
                            
                            # Get Default Value if available
                            default_value = "Not Available"
                            if param_name in enhanced_import_params:
                                default_value = enhanced_import_params[param_name]['Default Value']
                                
                            print(f"  - {param_name}: Found as {source_type} in {source_path}, Default Value: {default_value}")
                        else:
                            print(f"  - {param_name}: Not found in any import source")
                    
                    # Store enhanced parameter data for use in export
                    self.impl_scanner.enhanced_import_params = enhanced_import_params
                    
                    # Store for JSON processing
                    self.enhanced_params_map[main_path] = enhanced_import_params
                    self.import_sources_map[main_path] = normalized_import_paths
                
                # Process enumerations if any found
                if enum_elements:
                    print("\nProcessing enum elements...")
                    enum_data = self.process_class_enumerations(main_path)
                    if enum_data:
                        print(f"Found and processed {len(enum_data)} enumerations")
                    
                # Export the data
                self.export_class_data(main_path, normalized_import_paths)
                return True
                
            else:
                print(f"Main class not found: {main_class_path}")
                return False
            
        except Exception as e:
            print(f"Error processing cross-class import: {str(e)}")
            traceback.print_exc()
            return False

    def process_enumeration(self, enum_element, class_path, enum_name=None):
        """Process details of an enumeration element with enhanced detection
        
        Args:
            enum_element: The ASCET enumeration element or component
            class_path: Path to the class containing the enumeration
            enum_name: Optional name of the enumeration (if already known)
            
        Returns:
            Dictionary with enumeration details
        """
        try:
            if not enum_element:
                print(f"Invalid enumeration element")
                return None
                
            if enum_name is None:
                enum_name = enum_element.GetName() if hasattr(enum_element, 'GetName') else "Unknown"
                
            print(f"\nProcessing enumeration: {enum_name} from {class_path}")
            
            # Initialize enum data
            enum_data = {
                'name': enum_name,
                'path': class_path,
                'enumerators': [],
                'implementation': {}
            }
            
            # Get scope and kind if available
            if hasattr(enum_element, 'GetScope'):
                enum_data['scope'] = self.impl_scanner.get_scope(enum_element)
            
            if hasattr(enum_element, 'IsConstant') or hasattr(enum_element, 'IsParameter') or hasattr(enum_element, 'IsVariable'):
                enum_data['kind'] = self.impl_scanner.get_kind(enum_element)
            
            # Get comment
            if hasattr(enum_element, 'GetComment'):
                comment = enum_element.GetComment()
                if comment:
                    enum_data['comment'] = comment
            
            # ENHANCED: Check if the element is directly an enumeration component
            is_enum_component = hasattr(enum_element, 'IsEnumeration') and enum_element.IsEnumeration()
            
            # Try multiple approaches to get enumerators
            
            # Approach 1: If it's an enum component, try GetEnumerators directly
            if is_enum_component and hasattr(enum_element, 'GetEnumerators'):
                try:
                    enumerators = enum_element.GetEnumerators()
                    if enumerators:
                        enum_data['enumerators'] = enumerators
                        print(f"Found {len(enumerators)} enumerators from component")
                except Exception as ee:
                    print(f"Error getting enumerators from component: {str(ee)}")
            
            # Approach 2: Try to get the value object to extract enumerators
            if not enum_data['enumerators'] and hasattr(enum_element, 'GetValue'):
                value_obj = enum_element.GetValue()
                
                # Get enumerators from the value object
                if value_obj and hasattr(value_obj, 'GetEnumerators'):
                    try:
                        enumerators = value_obj.GetEnumerators()
                        if enumerators:
                            enum_data['enumerators'] = enumerators
                            print(f"Found {len(enumerators)} enumerators from value object")
                    except Exception as ee:
                        print(f"Error getting enumerators from value object: {str(ee)}")
            
            # Get implementation information
            if hasattr(enum_element, 'GetImplementation'):
                try:
                    impl = enum_element.GetImplementation()
                    if impl:
                        impl_data = {}
                        
                        # Get implementation type
                        if hasattr(impl, 'GetImplType'):
                            impl_type = impl.GetImplType()
                            if impl_type:
                                impl_data['type'] = impl_type
                        
                        # Try to get range information
                        if hasattr(impl, 'GetImplInfoForValue'):
                            impl_info = impl.GetImplInfoForValue()
                            if impl_info:
                                range_methods = [
                                    ('GetDoublePhysicalRange', float),
                                    ('GetFloatPhysicalRange', float),
                                    ('GetIntegerPhysicalRange', int),
                                    ('GetLongPhysicalRange', int),
                                    ('GetDoubleImplRange', float),
                                    ('GetFloatImplRange', float),
                                    ('GetIntegerImplRange', int),
                                    ('GetLongImplRange', int)
                                ]
                                
                                for method_name, convert_type in range_methods:
                                    if hasattr(impl_info, method_name):
                                        try:
                                            method = getattr(impl_info, method_name)
                                            range_value = method()
                                            if range_value and len(range_value) == 2:
                                                impl_data['min'] = convert_type(range_value[0])
                                                impl_data['max'] = convert_type(range_value[1])
                                                break
                                        except Exception:
                                            continue
                        
                        # Get max size
                        if hasattr(impl, 'GetMaxSize'):
                            try:
                                max_size = impl.GetMaxSize()
                                if max_size is not None:
                                    impl_data['max_size'] = max_size
                            except Exception:
                                pass
                        
                        if impl_data:
                            enum_data['implementation'] = impl_data
                except Exception as ie:
                    print(f"Error getting implementation info: {str(ie)}")
            
            return enum_data
            
        except Exception as e:
            print(f"Error processing enumeration: {str(e)}")
            traceback.print_exc()
            return None
    
    def process_class_enumerations(self, class_path):
        """Process all enumeration elements found in a class
        
        Args:
            class_path: Path to the class to process
            
        Returns:
            List of dictionaries with enumeration details
        """
        try:
            # First get enum elements from the class if not already done
            if class_path not in self.enum_elements:
                if not self.process_class(class_path):
                    print(f"Failed to process class {class_path}")
                    return []
            
            enum_elements = self.enum_elements.get(class_path, [])
            if not enum_elements:
                print(f"No enum elements found in class {class_path}")
                return []
            
            print(f"\nProcessing {len(enum_elements)} enum elements in class {class_path}")
            
            processed_enums = []
            for element_data in enum_elements:
                element_name = element_data['Name']
                print(f"Processing enum element: {element_name}")
                
                # Get the actual element from the class
                class_obj = self.impl_scanner.available_classes.get(class_path)
                if not class_obj:
                    print(f"Class object not found for {class_path}")
                    continue
                
                if hasattr(class_obj, 'GetModelElement'):
                    element = class_obj.GetModelElement(element_name)
                    if element:
                        enum_data = self.process_enumeration(element, class_path, element_name)
                        if enum_data:
                            processed_enums.append(enum_data)
                            # Store for future reference
                            key = f"{class_path}\\{element_name}"
                            self.enumeration_details[key] = enum_data
                    else:
                        print(f"Element {element_name} not found in class {class_path}")
                else:
                    print(f"Class {class_path} does not have GetModelElement method")
            
            print(f"Successfully processed {len(processed_enums)} enumerations from class {class_path}")
            return processed_enums
                
        except Exception as e:
            print(f"Error processing class enumerations: {str(e)}")
            traceback.print_exc()
            return []

    def export_enumerations_to_excel(self, enumerations, output_dir=None):
        """Export all enumeration data to a single Excel file with separate sheets
        
        Args:
            enumerations: List of enumeration dictionaries to export
            output_dir: Directory to save the Excel file
            
        Returns:
            Boolean indicating success or failure
        """
        try:
            if not enumerations:
                print("No enumeration information to export")
                return False
                
            if not output_dir:
                output_dir = "ASCET_Enumerations"
            os.makedirs(output_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = os.path.join(output_dir, f"Combined_Enumerations_{timestamp}.xlsx")
            
            # Create a single workbook for all enumerations
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                ws = wb["Sheet"]
                wb.remove(ws)
            
            # Define styles to be used across all sheets
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center')
            
            # Keep track of sheet names to ensure uniqueness
            used_sheet_names = set()
            
            # Create a summary sheet
            summary_sheet = wb.create_sheet(title="Summary")
            
            # Add headers to summary sheet
            summary_sheet['A1'] = "Enumeration Name"
            summary_sheet['B1'] = "Path"
            summary_sheet['C1'] = "Implementation Type"
            summary_sheet['D1'] = "Count of Values"
            
            # Style the header row
            for col in ['A', 'B', 'C', 'D']:
                summary_sheet[f'{col}1'].font = header_font
                summary_sheet[f'{col}1'].fill = header_fill
                summary_sheet[f'{col}1'].border = thin_border
                summary_sheet[f'{col}1'].alignment = center_alignment
            
            # Process each enumeration
            for i, enum_data in enumerate(enumerations, 2):  # Start from row 2 (after header)
                enum_name = enum_data['name']
                
                # Add to summary sheet
                summary_sheet[f'A{i}'] = enum_name
                summary_sheet[f'B{i}'] = enum_data['path']
                
                # Get implementation type if available
                impl_type = "N/A"
                impl = enum_data.get('implementation', {})
                if impl and 'type' in impl:
                    impl_type = impl['type']
                summary_sheet[f'C{i}'] = impl_type
                
                # Get count of values
                enum_count = len(enum_data.get('enumerators', []))
                summary_sheet[f'D{i}'] = enum_count
                
                # Style the data row
                for col in ['A', 'B', 'C', 'D']:
                    summary_sheet[f'{col}{i}'].border = Side(style='thin')
                
                # Create a safe sheet name
                safe_name = enum_name.replace('/', '_').replace('\\', '_').replace(':', '_')
                
                # Ensure sheet name is unique and not too long (Excel limit is 31 chars)
                if len(safe_name) > 28:
                    safe_name = safe_name[:28]
                
                # Handle duplicate sheet names
                original_name = safe_name
                suffix = 1
                while safe_name in used_sheet_names:
                    safe_name = f"{original_name[:28-len(str(suffix))]}_{suffix}"
                    suffix += 1
                    
                used_sheet_names.add(safe_name)
                
                # Create a sheet for this enumeration
                sheet = wb.create_sheet(title=safe_name)
                
                # Add title and metadata
                sheet['A1'] = f"Enumeration: {enum_name}"
                sheet['A1'].font = Font(bold=True, size=14)
                sheet.merge_cells('A1:D1')
                
                # Add path and other metadata
                sheet['A2'] = f"Path: {enum_data['path']}"
                sheet.merge_cells('A2:D2')
                
                row_idx = 3
                
                # Add scope and kind if available
                if 'scope' in enum_data:
                    sheet[f'A{row_idx}'] = f"Scope: {enum_data['scope']}"
                    row_idx += 1
                
                if 'kind' in enum_data:
                    sheet[f'A{row_idx}'] = f"Kind: {enum_data['kind']}"
                    row_idx += 1
                
                # Add implementation info if available
                if impl:
                    sheet[f'A{row_idx}'] = f"Implementation Type: {impl.get('type', 'N/A')}"
                    row_idx += 1
                    
                    if 'min' in impl and 'max' in impl:
                        sheet[f'A{row_idx}'] = f"Range: {impl['min']} to {impl['max']}"
                        row_idx += 1
                    
                    if 'max_size' in impl:
                        sheet[f'A{row_idx}'] = f"Max Size: {impl['max_size']}"
                        row_idx += 1
                
                # Add comment if available
                if 'comment' in enum_data and enum_data['comment']:
                    sheet[f'A{row_idx}'] = f"Comment: {enum_data['comment']}"
                    row_idx += 1
                
                # Add a blank row
                row_idx += 1
                
                # Horizontal layout for enumeration values
                if 'enumerators' in enum_data and enum_data['enumerators']:
                    # Add "Value" label in first column
                    sheet[f'A{row_idx}'] = "Value"
                    sheet[f'A{row_idx}'].font = header_font
                    sheet[f'A{row_idx}'].fill = header_fill
                    sheet[f'A{row_idx}'].border = thin_border
                    sheet[f'A{row_idx}'].alignment = center_alignment
                    
                    # Add Enumerator labels horizontally
                    row_idx += 1
                    sheet[f'A{row_idx}'] = "Enumerator"
                    sheet[f'A{row_idx}'].font = header_font
                    sheet[f'A{row_idx}'].fill = header_fill
                    sheet[f'A{row_idx}'].border = thin_border
                    sheet[f'A{row_idx}'].alignment = center_alignment
                    
                    # Fill in values horizontally
                    for j, enumerator in enumerate(enum_data['enumerators']):
                        col = get_column_letter(j + 2)  # Start from column B
                        
                        # Value in first row
                        value_cell = sheet[f'{col}{row_idx-1}']
                        value_cell.value = j
                        value_cell.border = thin_border
                        value_cell.alignment = center_alignment
                        
                        # Enumerator in second row
                        name_cell = sheet[f'{col}{row_idx}']
                        name_cell.value = enumerator
                        name_cell.border = thin_border
                        name_cell.alignment = center_alignment
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = None
                    
                    for cell in column:
                        # Skip cells that are part of a merged range
                        is_merged = False
                        for merged_range in sheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                is_merged = True
                                break
                        
                        if is_merged:
                            continue
                            
                        if column_letter is None:
                            column_letter = cell.column_letter
                            
                        try:
                            if cell.value is not None:
                                cell_str = str(cell.value)
                                if len(cell_str) > max_length:
                                    max_length = len(cell_str)
                        except:
                            pass
                    
                    if column_letter and max_length > 0:
                        adjusted_width = max(8, max_length + 2)  # Minimum width of 8 for better readability
                        sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Auto-adjust column widths for summary sheet
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = None
                
                for cell in column:
                    if column_letter is None:
                        column_letter = cell.column_letter
                        
                    try:
                        if cell.value is not None:
                            cell_str = str(cell.value)
                            if len(cell_str) > max_length:
                                max_length = len(cell_str)
                    except:
                        pass
                
                if column_letter and max_length > 0:
                    adjusted_width = max(15, max_length + 2)  # Wider minimum for summary
                    summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(excel_path)
            print(f"Successfully exported {len(enumerations)} enumerations to combined Excel file: {excel_path}")
            return True
            
        except Exception as e:
            print(f"Error exporting enumerations to Excel: {str(e)}")
            traceback.print_exc()
            return False
        
    # ========== JSON FUNCTIONALITY FROM ASCETDataInterface ==========
    
    def get_default_value_fixed(self, element):
        """
        修复的默认值获取方法，直接采用用户验证成功的简单方法
        Parameters:
            element: ASCET模型元素
        Returns:
            str: 默认值或'---'
        """
        try:
            # 直接使用与用户成功代码相同的逻辑
            di = element.GetValue()
            if not di:
                print(f"    GetValue() 返回 None")
                return '---'
                
            # 按常见类型优先级尝试（与用户成功代码完全相同的顺序）
            type_methods = [
                'GetDoubleValue',   # 双精度浮点
                'GetFloatValue',    # 单精度浮点  
                'GetIntegerValue',  # 整数
                'GetBooleanValue',  # 布尔
                'GetStringValue',   # 字符串
                'GetLongValue'      # 长整型
            ]
            
            for method_name in type_methods:
                if hasattr(di, method_name):
                    try:
                        value = getattr(di, method_name)()
                        if value is not None:
                            print(f"    ✓ 使用 {method_name} 获取到默认值: {value}")
                            return str(value)
                    except Exception as e:
                        print(f"    {method_name} 失败: {e}")
                        continue
            
            print(f"    ✗ 所有数据类型方法都失败")
            return '---'
            
        except Exception as e:
            print(f"    ✗ 获取默认值失败: {e}")
            return '---'

    def fix_log_default_values(self, class_path):
        """
        专门修复log类型元素的Default Value
        """
        try:
            print(f"\n{'='*60}")
            print(f"修复log类型Default Value: {class_path}")
            print(f"{'='*60}")
            
            # 获取类的实现数据
            all_elements = self.impl_scanner.all_implementations.get(class_path, [])
            if not all_elements:
                print("没有找到实现数据")
                return False
            
            # 获取类对象
            class_obj = self.impl_scanner.available_classes.get(class_path)
            if not class_obj:
                print("没有找到类对象")
                return False
            
            # 查找所有log类型且Default Value为'---'的元素
            log_elements_to_fix = []
            for element_data in all_elements:
                if (element_data.get('Type', '').lower() == 'log' and 
                    element_data.get('Default Value', '---') == '---' and
                    element_data.get('Scope') in ['Local', 'Imported'] and
                    element_data.get('Kind') in ['Parameter', 'Variable', 'Constant']):
                    log_elements_to_fix.append(element_data)
            
            print(f"找到{len(log_elements_to_fix)}个需要修复的log类型元素:")
            for elem in log_elements_to_fix:
                print(f"  - {elem['Name']} ({elem['Scope']} {elem['Kind']})")
            
            # 修复每个log类型元素的Default Value
            fixed_count = 0
            for element_data in log_elements_to_fix:
                element_name = element_data['Name']
                print(f"\n修复元素: {element_name}")
                
                if hasattr(class_obj, 'GetModelElement'):
                    element = class_obj.GetModelElement(element_name)
                    if element:
                        # 使用简单直接的方法获取默认值
                        try:
                            di = element.GetValue()
                            if di:
                                # 按优先级尝试不同数据类型方法
                                type_methods = [
                                    'GetDoubleValue', 'GetFloatValue', 'GetIntegerValue',
                                    'GetBooleanValue', 'GetStringValue', 'GetLongValue'
                                ]
                                
                                default_found = False
                                for method_name in type_methods:
                                    if hasattr(di, method_name):
                                        try:
                                            value = getattr(di, method_name)()
                                            if value is not None:
                                                element_data['Default Value'] = str(value)
                                                print(f"  ✅ 使用 {method_name} 获取到log默认值: {value}")
                                                fixed_count += 1
                                                default_found = True
                                                break
                                        except Exception as e:
                                            print(f"  {method_name} 失败: {e}")
                                            continue
                                
                                if not default_found:
                                    # 如果获取失败，log类型通常默认为0（false）
                                    element_data['Default Value'] = '0'
                                    print(f"  ✅ 使用log类型标准默认值: 0")
                                    fixed_count += 1
                            else:
                                # GetValue失败，使用标准默认值
                                element_data['Default Value'] = '0'
                                print(f"  ✅ GetValue失败，使用log类型标准默认值: 0")
                                fixed_count += 1
                        except Exception as e:
                            # 获取失败，使用标准默认值
                            element_data['Default Value'] = '0'
                            print(f"  ✅ 异常处理，使用log类型标准默认值: 0 (错误: {e})")
                            fixed_count += 1
                    else:
                        print(f"  ❌ 无法获取元素对象: {element_name}")
                else:
                    print(f"  ❌ 类对象不支持GetModelElement方法")
            
            print(f"\n修复完成: 成功修复 {fixed_count}/{len(log_elements_to_fix)} 个log类型元素")
            return fixed_count > 0
            
        except Exception as e:
            print(f"修复log类型Default Value时出错: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def collect_all_data(self, class_path: str) -> bool:
        """
        确保数据已收集并准备JSON导出（不重复处理）
        Args:
            class_path: Path to the class to process
        Returns:
            Boolean indicating success or failure
        """
        try:
            # Normalize path format
            normalized_path = self.normalize_class_path(class_path)
            
            # Check if we've already collected data for this class
            if normalized_path in self.cached_classes:
                print(f"Data for class {normalized_path} already collected and cached")
                return True
            
            # Check if the class exists in available classes
            best_match = self.find_best_matching_path(normalized_path)
            if not best_match:
                print(f"Class {normalized_path} not found in database")
                return False
            
            normalized_path = best_match
            
            # 检查是否已经有实现数据 - 如果有，说明已经处理过了
            if normalized_path in self.impl_scanner.all_implementations:
                print(f"Implementation data already exists for {normalized_path}, preparing JSON...")
                
                # Store enhanced parameters for JSON generation
                if hasattr(self.impl_scanner, 'enhanced_import_params'):
                    self.enhanced_params_map[normalized_path] = self.impl_scanner.enhanced_import_params.copy()
                
                # Mark as cached
                self.cached_classes.add(normalized_path)
                
                # Convert all collected data to JSON format
                self._prepare_json_data(normalized_path)
                
                return True
            
            # 如果没有数据，说明需要先处理
            print(f"No implementation data found for {normalized_path}, processing class first...")
            
            # Auto-discover import sources
            print("\nAuto-discovering import sources...")
            import_sources = self.discover_import_sources(normalized_path)
            if import_sources:
                print(f"Discovered {len(import_sources)} import sources for {normalized_path}")
                self.import_sources_map[normalized_path] = import_sources
            else:
                print(f"No import sources discovered for {normalized_path}")
            
            # 直接处理实现数据，不调用 process_cross_class_import 避免循环
            success = self._process_class_data_only(normalized_path, import_sources)
            
            if success:
                # Extract enumerations specifically for this class
                print(f"\nExtracting enumerations from class {normalized_path}...")
                self.extract_class_enumerations(normalized_path)
                
                # *** 关键修复：专门处理log类型的Default Value ***
                print(f"\nApplying final fix for log type Default Values...")
                self.fix_log_default_values(normalized_path)
                
                # Store enhanced parameters for later JSON generation
                if hasattr(self.impl_scanner, 'enhanced_import_params'):
                    self.enhanced_params_map[normalized_path] = self.impl_scanner.enhanced_import_params.copy()
                
                # Mark as cached
                self.cached_classes.add(normalized_path)
                
                # Convert all collected data to JSON format
                self._prepare_json_data(normalized_path)
                
                return True
            else:
                print(f"Failed to process class data for {normalized_path}")
                return False
            
        except Exception as e:
            print(f"Error collecting data for class {class_path}: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _process_class_data_only(self, main_path, import_class_paths=None):
        """
        只处理类数据，不导出（避免循环调用）
        Args:
            main_path: Path to the main class
            import_class_paths: Optional list of paths to import source classes
            
        Returns:
            Boolean indicating success or failure
        """
        try:
            print(f"Processing class data only (no export): {main_path}")
            
            # Convert to list if a single string is provided
            if isinstance(import_class_paths, str):
                import_class_paths = [import_class_paths]
                
            # Normalize import paths
            normalized_import_paths = []
            if import_class_paths:
                for import_path in import_class_paths:
                    path = import_path.replace('/', '\\')
                    if path.startswith('\\'):
                        path = path[1:]  # Remove leading backslash
                    normalized_import_paths.append(path)
            
            # Process the main class first
            if not self.impl_scanner.process_selected_class(main_path):
                print(f"Failed to process implementation data for class {main_path}")
                return False
            
            # Get all elements from the implementation data
            all_elements = self.impl_scanner.all_implementations.get(main_path, [])
            
            # Split into imported parameters and non-imported elements
            imported_params = [item for item in all_elements if item['Scope'] == 'Imported' and item['Kind'] == 'Parameter']
            non_imported_elements = [item for item in all_elements if item['Scope'] != 'Imported']
            
            # Identify enum type elements
            enum_elements = [item for item in all_elements if item['Type'].lower() == 'enum']
            if enum_elements:
                print(f"\nFound {len(enum_elements)} enum type elements in class {main_path}")
                self.enum_elements[main_path] = enum_elements
            
            all_found_tables = []
            
            # First scan the main class itself for lookup tables, but only for non-imported elements
            if non_imported_elements:
                print(f"\nScanning main class {main_path} for lookup tables (non-imported elements only)...")
                non_imported_names = [elem['Name'] for elem in non_imported_elements]
                main_class_tables = self.table_handler.get_multiple_tables(main_path, non_imported_names)
                
                if main_class_tables:
                    print(f"Found {len(main_class_tables)} lookup tables for non-imported elements in main class {main_path}")
                    all_found_tables.extend(main_class_tables)
            
            # Handle imported parameters
            if imported_params and normalized_import_paths:
                print(f"\nFound {len(imported_params)} imported parameters")
                param_names = [param['Name'] for param in imported_params]
                
                # Store parameter sources and enhanced parameter data
                enhanced_import_params = {}
                
                # Process each import source class
                for idx, import_path in enumerate(normalized_import_paths):
                    print(f"\nProcessing import source {idx+1}/{len(normalized_import_paths)}: {import_path}")
                    
                    # Find class in database
                    import_exists = import_path in self.impl_scanner.available_classes
                    if not import_exists:
                        # Try to find similar paths
                        matched_paths = [path for path in self.impl_scanner.available_classes.keys() 
                                    if import_path.lower() in path.lower()]
                        
                        if matched_paths:
                            import_path = matched_paths[0]
                            print(f"Using best match for import class: {import_path}")
                            import_exists = True
                        else:
                            print(f"Import class not found: {import_path}")
                            continue
                    
                    if import_exists:
                        print(f"Searching for parameters in {import_path}...")
                        
                        # Process the import class to get implementation data
                        if import_path not in self.impl_scanner.all_implementations:
                            self.impl_scanner.process_selected_class(import_path)
                        
                        # Get the implementation data
                        import_impl_data = self.impl_scanner.all_implementations.get(import_path, [])
                        
                        # Try to find lookup tables and implementation information for each parameter
                        for param_name in param_names:
                            if param_name in enhanced_import_params:
                                continue
                                
                            # First, try to get as a lookup table
                            table = self.table_handler.get_table_by_name(import_path, param_name)
                            
                            if table:
                                print(f"Found lookup table for parameter: {param_name}")
                                all_found_tables.append(table)
                            
                            # Find implementation data for this parameter
                            param_impl = next((item for item in import_impl_data if item['Name'] == param_name), None)
                            if param_impl:
                                # Make sure we have the Default Value properly set 
                                if param_impl['Default Value'] == '---' and param_impl['Q'] != '---':
                                    param_impl['Default Value'] = param_impl['Q']
                                
                                enhanced_import_params[param_name] = param_impl
                                print(f"  - Default Value for {param_name}: {param_impl['Default Value']}")
                
                # Store enhanced parameter data
                self.impl_scanner.enhanced_import_params = enhanced_import_params
            
            # Store all found tables
            if all_found_tables:
                print(f"\nFound {len(all_found_tables)} lookup tables in total")
                self.all_tables[main_path] = all_found_tables
            
            return True
            
        except Exception as e:
            print(f"Error processing class data only: {str(e)}")
            traceback.print_exc()
            return False
    
    def _prepare_json_data(self, class_path: str) -> None:
        """
        Prepare JSON data for a specific class from collected data
        Args:
            class_path: Path to the class
        """
        # Create a master dictionary for this class
        class_data = {
            "class_path": class_path,
            "timestamp": datetime.now().isoformat(),
            "import_sources": self.import_sources_map.get(class_path, []),
            "signals": []
        }
        
        # Get enhanced parameters for this class
        enhanced_params = self.enhanced_params_map.get(class_path, {})
        
        # Add implementation data
        impl_data = self.impl_scanner.all_implementations.get(class_path, [])
        for item in impl_data:
            signal = self._convert_implementation_to_json(item)
            signal["class_path"] = class_path
            
            # Handle imported parameters - apply enhanced data from source classes
            if item['Scope'] == 'Imported' and item['Kind'] == 'Parameter':
                param_name = item['Name']
                
                # Check if we have enhanced data for this parameter
                if param_name in enhanced_params:
                    enhanced_data = enhanced_params[param_name]
                    
                    # Update signal with enhanced data
                    for key, value in enhanced_data.items():
                        if key in signal and signal[key] == "---" and value != "---":
                            signal[key] = value
                    
                    
                
                # Find source information
                source_info = self._find_parameter_source(class_path, param_name)
                if source_info:
                    signal["import_source"] = source_info
                    
                    # *** 新增逻辑：使用import_source中的calibration数据更新主信号的Calibration字段 ***
                    if "calibration" in source_info and source_info["calibration"] != "---":
                        # 更新主信号的Calibration字段为import_source中的calibration值
                        signal["Calibration"] = source_info["calibration"]
                        print(f"Updated Calibration for imported parameter '{param_name}': {source_info['calibration']}")
                    
                    # 也可以根据需要设置其他字段
                    # 例如，如果源类中有更准确的Min/Max值
                    if "min" in source_info and signal.get("Min", "---") == "---":
                        signal["Min"] = source_info["min"]
                    if "max" in source_info and signal.get("Max", "---") == "---":
                        signal["Max"] = source_info["max"]
            
            # Check if this signal has a lookup table
            table = self._find_lookup_table(class_path, item['Name'])
            if table:
                # Convert the table to JSON and check if it's valid
                table_json = self._convert_table_to_json(table)
                
                # Only include the lookup table if it's valid (has meaningful data)
                if self._is_valid_lookup_table(table_json):
                    signal["lookup_table"] = table_json
            
            # Check if this signal is an enumeration
            if item['Type'].lower() == 'enum':
                enum_data = self._find_enumeration(class_path, item['Name'])
                if enum_data:
                    signal["enumeration"] = self._convert_enumeration_to_json(enum_data)
            
            class_data["signals"].append(signal)
        
        # Cache the prepared JSON data
        self.json_cache[class_path] = class_data
    
    def _is_valid_lookup_table(self, table_json: Dict) -> bool:
        """
        Check if a lookup table has valid data that should be included in JSON
        Args:
            table_json: Lookup table in JSON format
        Returns:
            Boolean indicating if the table has meaningful data
        """
        table_type = table_json.get("type", "")
        
        # For 1D tables
        if table_type == "1D":
            # Check size
            size = table_json.get("size", 0)
            if size is None or size <= 1:
                # Also check if there are actual values
                x_values = table_json.get("x_values", [])
                values = table_json.get("values", [])
                
                # If both x_values and values are empty or have less than 2 points, consider invalid
                if len(x_values) <= 1 and len(values) <= 1:
                    return False
        
        # For 2D tables
        elif table_type == "2D":
            x_values = table_json.get("x_values", [])
            y_values = table_json.get("y_values", [])
            values = table_json.get("values", [])
            
            # If x_values or y_values are empty, consider invalid
            if not x_values or not y_values or not values:
                return False
            
            # If values is empty or doesn't match expected size, consider invalid
            x_size = len(x_values)
            y_size = len(y_values)
            
            if y_size > 0 and isinstance(values, list):
                # Check if values matrix has expected dimensions
                if len(values) != y_size:
                    return False
                
                # For 2D table, we should have at least one non-empty row
                has_valid_row = False
                for row in values:
                    if isinstance(row, list) and len(row) > 0:
                        has_valid_row = True
                        break
                
                if not has_valid_row:
                    return False
        
        return True
    
    def _find_parameter_source(self, class_path: str, param_name: str) -> Dict:
        """
        Find the source class for an imported parameter with enhanced details
        Args:
            class_path: Path to the class
            param_name: Name of the parameter
        Returns:
            Dictionary with source information or empty dict if not found
        """
        import_sources = self.import_sources_map.get(class_path, [])
        
        for source_path in import_sources:
            # Check if parameter exists in this source
            if source_path in self.impl_scanner.all_implementations:
                source_elements = self.impl_scanner.all_implementations[source_path]
                for element in source_elements:
                    if element['Name'] == param_name:
                        source_info = {
                            "source_class": source_path,
                            "type": element.get('Type', '---'),
                            "unit": element.get('Unit', '---'),
                            "comment": element.get('Comment', '---'),
                            "calibration": element.get('Calibration', '---')
                        }
                        
                        # Add Min/Max values if available
                        if element.get('Min', '---') != '---':
                            source_info['min'] = element['Min']
                        if element.get('Max', '---') != '---':
                            source_info['max'] = element['Max']
                        
                        # Add Q and Default Value if available
                        if element.get('Q', '---') != '---':
                            source_info['q'] = element['Q']
                        
                        default_value = element.get('Default Value', '---')
                        if default_value == '---' and element.get('Q', '---') != '---':
                            default_value = element['Q']
                        
                        if default_value != '---':
                            source_info['default_value'] = default_value
                        
                        return source_info
        
        return {}
    
    def _convert_implementation_to_json(self, impl_item: Dict) -> Dict:
        """
        Convert implementation item to JSON format
        Args:
            impl_item: Implementation item dictionary
        Returns:
            JSON compatible dictionary
        """
        # Create a copy to avoid modifying the original
        signal = impl_item.copy()
        
        # Convert any non-JSON serializable values
        for key, value in signal.items():
            if isinstance(value, (datetime, set)):
                signal[key] = str(value)
            elif isinstance(value, (complex, type(None))):
                signal[key] = str(value)
        
        return signal
    
    def _convert_table_to_json(self, table: Dict) -> Dict:
        """
        Convert lookup table to JSON format
        Args:
            table: Lookup table dictionary
        Returns:
            JSON compatible dictionary
        """
        # Create a clean copy without potential non-serializable objects
        table_json = {
            "name": table.get("name", ""),
            "type": table.get("type", ""),
        }
        
        # Handle 1D tables
        if table.get("type") == "1D":
            # Convert x_values and values to serializable format
            x_values = table.get("x_values", [])
            values = table.get("values", [])
            
            # Ensure all values are JSON serializable
            table_json["x_values"] = [float(x) if isinstance(x, (int, float)) else str(x) for x in x_values]
            table_json["values"] = [float(v) if isinstance(v, (int, float)) else str(v) for v in values]
            table_json["size"] = table.get("size", table.get("max_size", 0))
            
            if "interpolation" in table:
                table_json["interpolation"] = table["interpolation"]
            if "extrapolation" in table:
                table_json["extrapolation"] = table["extrapolation"]
        
        # Handle 2D tables
        elif table.get("type") == "2D":
            # Convert x_values and y_values to serializable format
            x_values = table.get("x_values", [])
            y_values = table.get("y_values", [])
            
            # Ensure all values are JSON serializable
            table_json["x_values"] = [float(x) if isinstance(x, (int, float)) else str(x) for x in x_values]
            table_json["y_values"] = [float(y) if isinstance(y, (int, float)) else str(y) for y in y_values]
            
            # Handle matrix values with nested lists
            matrix = table.get("values", [])
            matrix_json = []
            
            for row in matrix:
                if isinstance(row, (list, tuple)):
                    # Convert each cell value to a serializable format
                    row_json = [float(cell) if isinstance(cell, (int, float)) else str(cell) for cell in row]
                    matrix_json.append(row_json)
                else:
                    # Handle single value
                    matrix_json.append(float(row) if isinstance(row, (int, float)) else str(row))
            
            table_json["values"] = matrix_json
            table_json["x_size"] = table.get("x_size", len(table.get("x_values", [])))
            table_json["y_size"] = table.get("y_size", len(table.get("y_values", [])))
        
        return table_json
    
    def _convert_enumeration_to_json(self, enum_data: Dict) -> Dict:
        """
        Convert enumeration data to JSON format
        Args:
            enum_data: Enumeration data dictionary
        Returns:
            JSON compatible dictionary
        """
        enum_json = {
            "name": enum_data.get("name", ""),
            "path": enum_data.get("path", ""),
            "enumerators": enum_data.get("enumerators", []),
        }
        
        # Add implementation details if available
        impl = enum_data.get("implementation", {})
        if impl:
            enum_json["implementation"] = {
                "type": impl.get("type", ""),
                "min": impl.get("min", 0),
                "max": impl.get("max", 0),
                "max_size": impl.get("max_size", 0)
            }
        
        # Add scope, kind, and comment if available
        if "scope" in enum_data:
            enum_json["scope"] = enum_data["scope"]
        if "kind" in enum_data:
            enum_json["kind"] = enum_data["kind"]
        if "comment" in enum_data and enum_data["comment"]:
            enum_json["comment"] = enum_data["comment"]
        
        return enum_json
    
    def _find_lookup_table(self, class_path: str, signal_name: str) -> Optional[Dict]:
        """
        Find lookup table for a signal in a class or its import sources
        Args:
            class_path: Path to the class
            signal_name: Name of the signal
        Returns:
            Lookup table dictionary or None if not found
        """
        # First check in main class
        tables = self.all_tables.get(class_path, [])
        for table in tables:
            if table.get("name") == signal_name:
                return table
        
        # If not found in main class, check import sources
        import_sources = self.import_sources_map.get(class_path, [])
        for source_path in import_sources:
            source_tables = self.all_tables.get(source_path, [])
            for table in source_tables:
                if table.get("name") == signal_name:
                    return table
        
        return None
    
    def _find_enumeration(self, class_path: str, signal_name: str) -> Optional[Dict]:
        """
        Find enumeration data for a signal in a class or referenced class
        Args:
            class_path: Path to the class
            signal_name: Name of the signal
        Returns:
            Enumeration data dictionary or None if not found
        """
        # First check in current class
        key = f"{class_path}\\{signal_name}"
        if key in self.enumeration_details:
            return self.enumeration_details[key]
        
        # Check all enumeration details to find by name
        for enum_key, enum_data in self.enumeration_details.items():
            if enum_data.get("name") == signal_name:
                return enum_data
        
        return None
    
    def get_json_data(self, class_path: str, indent: Optional[int] = None) -> str:
        """
        Get JSON data for a specific class
        Args:
            class_path: Path to the class
            indent: Optional indentation for pretty printing
        Returns:
            JSON string representation of the data
        """
        # Normalize path format
        normalized_path = self.normalize_class_path(class_path)
        
        # Find best matching path
        best_match = self.find_best_matching_path(normalized_path)
        if not best_match:
            return json.dumps({"error": f"Class {normalized_path} not found"})
        
        normalized_path = best_match
        
        # Check if we need to collect data first
        if normalized_path not in self.cached_classes:
            if not self.collect_all_data(normalized_path):
                return json.dumps({"error": f"Failed to collect data for {normalized_path}"})
        
        # Return the cached JSON data
        return json.dumps(self.json_cache.get(normalized_path, {}), indent=indent)
    
    def get_signal_json_data(self, class_path: str, signal_name: str, indent: Optional[int] = None) -> str:
        """
        Get JSON data for a specific signal in a class
        Args:
            class_path: Path to the class
            signal_name: Name of the signal
            indent: Optional indentation for pretty printing
        Returns:
            JSON string representation of the signal data
        """
        # Normalize path format
        normalized_path = self.normalize_class_path(class_path)
        
        # Find best matching path
        best_match = self.find_best_matching_path(normalized_path)
        if not best_match:
            return json.dumps({"error": f"Class {normalized_path} not found"})
        
        normalized_path = best_match
        
        # Check if we need to collect data first
        if normalized_path not in self.cached_classes:
            if not self.collect_all_data(normalized_path):
                return json.dumps({"error": f"Failed to collect data for {normalized_path}"})
        
        # Find the specific signal
        class_data = self.json_cache.get(normalized_path, {})
        signals = class_data.get("signals", [])
        
        for signal in signals:
            if signal.get("Name") == signal_name:
                return json.dumps(signal, indent=indent)
        
        return json.dumps({"error": f"Signal {signal_name} not found in class {normalized_path}"})
    
    def save_json_to_file(self, class_path: str, output_file: Optional[str] = None, indent: int = 4) -> bool:
        """
        Save JSON data for a class to a file
        Args:
            class_path: Path to the class
            output_file: Optional file path for output
            indent: Indentation for pretty printing
        Returns:
            Boolean indicating success or failure
        """
        try:
            # Normalize path format
            normalized_path = self.normalize_class_path(class_path)
            
            # Find best matching path
            best_match = self.find_best_matching_path(normalized_path)
            if not best_match:
                print(f"Class {normalized_path} not found")
                return False
            
            normalized_path = best_match
            
            # Check if we need to collect data first
            if normalized_path not in self.cached_classes:
                if not self.collect_all_data(normalized_path):
                    print(f"Failed to collect data for {normalized_path}")
                    return False
            
            # Create default output file name if not provided
            if not output_file:
                base_name = os.path.basename(normalized_path)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"{base_name}_ASCET_Data_{timestamp}.json"
            
            # Create directory if it doesn't exist
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # Write the JSON data to file
            with open(output_file, 'w') as f:
                json.dump(self.json_cache.get(normalized_path, {}), f, indent=indent)
            
            print(f"Successfully saved JSON data to {output_file}")
            return True
            
        except Exception as e:
            print(f"Error saving JSON data to file: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_available_classes_json(self, indent: Optional[int] = None) -> str:
        """
        Get JSON representation of all available classes
        Args:
            indent: Optional indentation for pretty printing
        Returns:
            JSON string with all available classes
        """
        classes_data = {
            "timestamp": datetime.now().isoformat(),
            "available_classes": list(self.impl_scanner.available_classes.keys())
        }
        
        return json.dumps(classes_data, indent=indent)
    
    def get_all_cached_data_json(self, indent: Optional[int] = None) -> str:
        """
        Get JSON representation of all cached data
        Args:
            indent: Optional indentation for pretty printing
        Returns:
            JSON string with all cached data
        """
        all_data = {
            "timestamp": datetime.now().isoformat(),
            "cached_classes": list(self.cached_classes),
            "class_data": self.json_cache
        }
        
        return json.dumps(all_data, indent=indent)
        
    def disconnect(self):
        """Disconnect from ASCET"""
        self.impl_scanner.disconnect()
        # No need to disconnect table_handler since it shares the connection

def main():
    """Main entry point for the enhanced ASCET scanner with automatic import source discovery and JSON export"""
    # Define ASCET version
    ascet_version = "6.1.4"
    
    # Define only the main class path - import sources will be auto-discovered
    main_class_path = r"\Customer\CC_CN\Package\ECAS_ElectronicallyControlledAirSpring\private\ECAS_HC_Axlemoving4DZSelection"
    
    # Initialize the integrated scanner
    scanner = IntegratedAscetScanner(ascet_version)
    
    try:
        # Connect to ASCET
        if not scanner.connect():
            print("Failed to connect to ASCET. Exiting.")
            return
        
        # Scan database structure
        print("\nScanning ASCET database structure...")
        if not scanner.scan_database_structure():
            print("Failed to scan database structure. Exiting.")
            return
        
        # Process the main class without specifying import sources
        # Import sources will be auto-discovered and both Excel and JSON will be exported
        print("\nProcessing class with auto-discovery of import sources...")
        success = scanner.process_class(main_class_path)  # This will handle both processing and exporting
        
        if success:
            print("\n" + "="*80)
            print("✅ PROCESSING COMPLETED SUCCESSFULLY!")
            print("📊 Excel files exported: Implementation data, Lookup tables, Enumerations")
            print("📄 JSON file exported: Complete structured data")
            print("="*80)
        else:
            print("\n❌ Processing failed!")
        
    except Exception as e:
        print(f"\nError during processing: {str(e)}")
        import traceback
        traceback.print_exc()
        
    finally:
        # Disconnect from ASCET
        scanner.disconnect()
        print("\nASCET processing completed.")

if __name__ == '__main__':
    main()